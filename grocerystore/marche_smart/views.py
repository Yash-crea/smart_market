from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.db import models
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from .models import Product, SmartProducts, Cart, CartItem, Order, OrderItem, Payment, Notification, MLForecastModel, UserProfile
from django.db.models import Sum, Avg, Max, Count, Q, F
from decimal import Decimal
import uuid
import json
import random
import logging
from datetime import datetime, timedelta

logger = logging.getLogger(__name__)

# ============= ML RECOMMENDATION FUNCTIONS =============

def get_current_season():
    """Get current season from current month"""
    current_month = datetime.now().month
    season_map = {
        12: 'winter', 1: 'winter', 2: 'winter',
        3: 'spring', 4: 'spring', 5: 'spring',
        6: 'summer', 7: 'summer', 8: 'summer',
        9: 'monsoon', 10: 'monsoon', 11: 'monsoon'
    }
    return season_map.get(current_month, 'all_year')

def get_upcoming_festivals():
    """Get festivals in next 30 days"""
    current_month = datetime.now().month
    festivals = []
    
    if current_month == 10:  # October
        festivals.append({'name': 'diwali', 'display_name': 'Diwali', 'date': '2026-10-24'})
    elif current_month in [3, 4]:  # March-April  
        festivals.append({'name': 'easter', 'display_name': 'Easter', 'date': '2026-04-05'})
    elif current_month == 12:  # December
        festivals.append({'name': 'christmas', 'display_name': 'Christmas', 'date': '2026-12-25'})
    elif current_month == 1:  # January
        festivals.append({'name': 'new_year', 'display_name': 'New Year', 'date': '2026-01-01'})
    
    return festivals


_last_demand_recalc = None

def recalculate_demand_scores():
    """
    Recalculate predicted_demand_7d and avg_weekly_sales from real OrderItem data.
    Throttled to run at most once every 10 minutes to avoid overhead on every page load.
    """
    global _last_demand_recalc
    now = datetime.now()

    # Throttle: skip if recalculated within the last 10 minutes
    if _last_demand_recalc and (now - _last_demand_recalc).total_seconds() < 600:
        return 0

    seven_days_ago = now - timedelta(days=7)
    thirty_days_ago = now - timedelta(days=30)

    # Count units sold per SmartProduct in last 7 days
    demand_7d = (
        OrderItem.objects
        .filter(smart_product__isnull=False, order__created_at__gte=seven_days_ago)
        .values('smart_product_id')
        .annotate(total_sold=Sum('quantity'))
    )
    demand_map_7d = {row['smart_product_id']: row['total_sold'] for row in demand_7d}

    # Average weekly sales from last 30 days
    demand_30d = (
        OrderItem.objects
        .filter(smart_product__isnull=False, order__created_at__gte=thirty_days_ago)
        .values('smart_product_id')
        .annotate(total_sold=Sum('quantity'))
    )
    demand_map_30d = {row['smart_product_id']: row['total_sold'] for row in demand_30d}

    updated = 0
    for sp in SmartProducts.objects.all():
        new_demand_7d = demand_map_7d.get(sp.id, 0)
        weeks_in_period = max(4.0, 1.0)  # 30 days ≈ 4 weeks
        new_avg_weekly = Decimal(str(round(demand_map_30d.get(sp.id, 0) / weeks_in_period, 2)))

        if sp.predicted_demand_7d != new_demand_7d or sp.avg_weekly_sales != new_avg_weekly:
            sp.predicted_demand_7d = new_demand_7d
            sp.avg_weekly_sales = new_avg_weekly
            sp.save(update_fields=['predicted_demand_7d', 'avg_weekly_sales'])
            updated += 1

    logger.info(f"Demand recalculation: updated {updated} products from {OrderItem.objects.count()} order items")
    _last_demand_recalc = now
    return updated


def get_ml_recommendations(user=None, algorithm='hybrid_ml', limit=10):
    """
    Get ML-powered recommendations using the ContextualRecommendationEngine.
    Returns SmartProduct instances with recommendation_reason and recommendation_type attached.
    Falls back to simple DB queries if ML engine fails.
    """
    try:
        from .ml_recommendations import ContextualRecommendationEngine

        engine = ContextualRecommendationEngine()
        result = engine.get_personalized_recommendations(
            user=user,
            algorithm=algorithm,
            limit=limit
        )

        recs = result.get('recommendations', [])
        if not recs:
            return []

        # Fetch actual product instances by ID
        rec_ids = [r['id'] for r in recs]
        # Build lookup: the engine may return Product or SmartProduct ids
        smart_lookup = {sp.id: sp for sp in SmartProducts.objects.filter(id__in=rec_ids)}
        product_lookup = {p.id: p for p in Product.objects.filter(id__in=rec_ids)}

        products = []
        for rec in recs:
            pid = rec['id']
            product = smart_lookup.get(pid) or product_lookup.get(pid)
            if product:
                reasons = rec.get('reasons', [])
                product.recommendation_reason = reasons[0] if reasons else rec.get('algorithm', 'ML recommended')
                product.ml_confidence = rec.get('confidence_score', rec.get('contextual_score', rec.get('hybrid_score', 0)))
                product.ml_reasons = reasons

                # Map algorithm source to recommendation_type
                if 'seasonal' in algorithm or 'contextual_score' in rec:
                    product.recommendation_type = 'seasonal'
                elif 'weather' in algorithm or 'weather_relevance_score' in rec:
                    product.recommendation_type = 'weather'
                elif 'behavior' in algorithm or 'behavioral_score' in rec:
                    product.recommendation_type = 'personal'
                else:
                    # Hybrid — use scoring factors to pick type
                    factors = rec.get('scoring_factors', [])
                    if 'user_behavior' in factors:
                        product.recommendation_type = 'personal'
                    elif 'weather_context' in factors:
                        product.recommendation_type = 'weather'
                    else:
                        product.recommendation_type = 'seasonal'
                products.append(product)

        return products
    except Exception as e:
        logger.warning(f"ML engine failed, using fallback: {e}")
        return []


def get_user_personal_recommendations(user, limit=4):
    """
    Recommend products based on the logged-in user's actual purchase history.
    Finds categories the user buys most, then recommends other products in those categories
    that the user hasn't purchased yet.
    """
    if not user or not user.is_authenticated:
        return []

    # Get user's purchased product IDs
    purchased_ids = set(
        OrderItem.objects
        .filter(order__user=user, smart_product__isnull=False)
        .values_list('smart_product_id', flat=True)
    )

    if not purchased_ids:
        return []

    # Find user's preferred categories (by order count)
    category_prefs = (
        OrderItem.objects
        .filter(order__user=user, smart_product__isnull=False)
        .values('smart_product__category')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')
    )
    preferred_cats = [row['smart_product__category'] for row in category_prefs if row['smart_product__category']]

    if not preferred_cats:
        return []

    # Recommend unseen products in preferred categories, ordered by demand
    personal_recs = list(
        SmartProducts.objects
        .filter(category__in=preferred_cats, stock_quantity__gt=0)
        .exclude(id__in=purchased_ids)
        .order_by('-predicted_demand_7d', '-avg_weekly_sales')[:limit]
    )

    for product in personal_recs:
        product.recommendation_reason = f"Based on your purchase history"
        product.recommendation_type = "personal"

    return personal_recs


def get_seasonal_recommendations(limit=6, user=None):
    """Get ML-powered seasonal recommendations with fallback to DB queries"""
    # First: recalculate demand from real order data
    recalculate_demand_scores()

    # Try ML engine first
    ml_recs = get_ml_recommendations(user=user, algorithm='ml_seasonal', limit=limit)
    if ml_recs:
        return ml_recs

    # Fallback: enhanced DB-based recommendations
    current_season = get_current_season()
    is_weekend = datetime.now().weekday() >= 5
    upcoming_festivals = get_upcoming_festivals()
    
    recommendations = []
    
    # 1. Seasonal Products (current season) — now ordered by REAL demand data
    seasonal_products = SmartProducts.objects.filter(
        peak_season=current_season,
        stock_quantity__gt=0
    ).order_by('-predicted_demand_7d', '-seasonal_priority')[:3]
    
    for product in seasonal_products:
        product.recommendation_reason = f"Perfect for {current_season} season"
        product.recommendation_type = "seasonal"
    
    recommendations.extend(seasonal_products)
    
    # 2. Weekend Favorites (if weekend)
    if is_weekend:
        weekend_favorites = SmartProducts.objects.filter(
            weekend_boost=True,
            stock_quantity__gt=0
        ).exclude(id__in=[p.id for p in recommendations]).order_by('-weekend_sales_multiplier')[:2]
        
        for product in weekend_favorites:
            product.recommendation_reason = "Weekend favorite"
            product.recommendation_type = "weekend"
        
        recommendations.extend(weekend_favorites)
    
    # 3. Festival Recommendations
    for festival in upcoming_festivals:
        festival_products = SmartProducts.objects.filter(
            festival_association=festival['name'],
            stock_quantity__gt=0
        ).exclude(id__in=[p.id for p in recommendations]).order_by('-festival_sales_boost')[:2]
        
        for product in festival_products:
            product.recommendation_reason = f"Great for {festival['display_name']}"
            product.recommendation_type = "festival"
        
        recommendations.extend(festival_products)
    
    # 4. Trending — now based on REAL order volume
    if len(recommendations) < limit:
        trending_products = SmartProducts.objects.filter(
            predicted_demand_7d__gt=0,
            stock_quantity__gt=0
        ).exclude(id__in=[p.id for p in recommendations]).order_by('-predicted_demand_7d')[:limit-len(recommendations)]
        
        for product in trending_products:
            product.recommendation_reason = f"Trending — {product.predicted_demand_7d} sold this week"
            product.recommendation_type = "trending"
        
        recommendations.extend(trending_products)
    
    # 5. Fill remaining slots with best sellers
    if len(recommendations) < limit:
        popular_products = SmartProducts.objects.filter(
            stock_quantity__gt=0
        ).exclude(id__in=[p.id for p in recommendations]).order_by('-avg_weekly_sales')[:limit-len(recommendations)]
        
        for product in popular_products:
            product.recommendation_reason = "Customer favorite"
            product.recommendation_type = "popular"
        
        recommendations.extend(popular_products)
    
    return recommendations[:limit]

def get_weather_based_recommendations(user=None):
    """Get weather-based product recommendations using ML engine with fallback"""
    # Try ML engine first
    ml_recs = get_ml_recommendations(user=user, algorithm='ml_weather', limit=3)
    if ml_recs:
        return ml_recs

    # Fallback: DB-based weather recommendations
    from .models import WeatherData
    today = datetime.now().date()
    
    try:
        try:
            weather = WeatherData.objects.get(date=today)
        except WeatherData.DoesNotExist:
            weather = WeatherData.objects.order_by('-date').first()
            if weather is None:
                return []

        weather_products = []

        if weather.condition in ['rainy', 'stormy']:
            weather_products = list(SmartProducts.objects.filter(
                weather_dependent=True,
                stock_quantity__gt=0
            ).filter(
                Q(name__icontains='tea') | Q(name__icontains='coffee') |
                Q(name__icontains='soup') | Q(name__icontains='chocolate')
            )[:3])
            reason = f"Perfect for {weather.condition} weather"

        elif weather.temperature_avg and weather.temperature_avg > 30:
            weather_products = list(SmartProducts.objects.filter(
                weather_dependent=True,
                stock_quantity__gt=0
            ).filter(
                Q(name__icontains='cold') | Q(name__icontains='ice') |
                Q(name__icontains='juice') | Q(name__icontains='water')
            )[:3])
            reason = f"Stay cool on hot days ({weather.temperature_avg}°C)"

        else:
            reason = "Weather appropriate"

        if not weather_products:
            weather_products = list(SmartProducts.objects.filter(
                weather_dependent=True,
                stock_quantity__gt=0
            ).order_by('-predicted_demand_7d')[:3])
            if weather_products:
                reason = f"Picks for today's weather ({weather.condition or 'current'})"

        for product in weather_products:
            product.recommendation_reason = reason
            product.recommendation_type = "weather"
            
        return weather_products
    except Exception:
        return []

# ============= END ML RECOMMENDATION FUNCTIONS =============

# ============= NOTIFICATION HELPER FUNCTIONS =============

def create_notification(notification_type, title, message, related_order=None, recipient_groups=['Owner']):
    """Create notifications for specific user groups"""
    recipients = User.objects.filter(groups__name__in=recipient_groups).distinct()
    
    for recipient in recipients:
        Notification.objects.create(
            recipient_user=recipient,
            notification_type=notification_type,
            title=title,
            message=message,
            related_order=related_order
        )

def get_unread_notifications_count(user):
    """Get count of unread notifications for user"""
    return Notification.objects.filter(recipient_user=user, is_read=False).count()

# ============= END NOTIFICATION HELPERS =============

# ============= GUEST CART HELPER FUNCTIONS =============

def get_session_cart(request):
    """Get cart from session for anonymous users"""
    return request.session.get('guest_cart', {})

def save_session_cart(request, cart_data):
    """Save cart to session for anonymous users"""
    request.session['guest_cart'] = cart_data
    request.session.modified = True

def add_to_session_cart(request, product_id, product_type, quantity=1):
    """Add item to session cart for anonymous users"""
    cart = get_session_cart(request)
    item_key = f"{product_type}_{product_id}"
    
    if item_key in cart:
        cart[item_key]['quantity'] += quantity
    else:
        cart[item_key] = {
            'product_id': product_id,
            'product_type': product_type,
            'quantity': quantity
        }
    
    save_session_cart(request, cart)
    return cart

def get_session_cart_items(request):
    """Get formatted cart items from session"""
    cart = get_session_cart(request)
    items = []
    total_items = 0
    total_amount = Decimal('0.00')
    
    for item_key, item_data in cart.items():
        try:
            if item_data['product_type'] == 'smart':
                product = SmartProducts.objects.get(id=item_data['product_id'])
                product_obj = None
                smart_product_obj = product
            else:
                product = Product.objects.get(id=item_data['product_id'])
                product_obj = product
                smart_product_obj = None
            
            subtotal = product.price * item_data['quantity']
            items.append({
                'product': product_obj,
                'smart_product': smart_product_obj,
                'quantity': item_data['quantity'],
                'subtotal': subtotal,
                'product_type': item_data['product_type'],
                'item_key': item_key,
                'product_name': product.name,
                'unit_price': product.price
            })
            total_items += item_data['quantity']
            total_amount += subtotal
        except (Product.DoesNotExist, SmartProducts.DoesNotExist):
            continue
    
    return items, total_items, total_amount

def merge_session_cart_to_user(request, user):
    """Merge session cart to user cart after login"""
    session_cart = get_session_cart(request)
    if not session_cart:
        return
    
    user_cart, created = Cart.objects.get_or_create(user=user)
    
    for item_key, item_data in session_cart.items():
        try:
            if item_data['product_type'] == 'smart':
                product = SmartProducts.objects.get(id=item_data['product_id'])
                cart_item, item_created = CartItem.objects.get_or_create(
                    cart=user_cart, smart_product=product,
                    defaults={'quantity': item_data['quantity']}
                )
            else:
                product = Product.objects.get(id=item_data['product_id'])
                cart_item, item_created = CartItem.objects.get_or_create(
                    cart=user_cart, product=product,
                    defaults={'quantity': item_data['quantity']}
                )
            
            if not item_created:
                cart_item.quantity += item_data['quantity']
                cart_item.save()
                
        except (Product.DoesNotExist, SmartProducts.DoesNotExist):
            continue
    
    # Clear session cart after merge
    request.session['guest_cart'] = {}
    request.session.modified = True

def clear_session_cart(request):
    """Clear session cart"""
    request.session['guest_cart'] = {}
    request.session.modified = True

# ============= END GUEST CART HELPERS =============

def unified_login(request):
    """Enhanced unified login for customers, owners, and staff"""
    if request.method == 'POST':
        username = request.POST.get('username') or request.POST.get('email')
        password = request.POST.get('password')
        
        if not username or not password:
            messages.error(request, 'Please provide both username/email and password.')
            return render(request, 'unified_login.html')
        
        # Try to authenticate user
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if not user.is_active:
                messages.error(request, 'Your account has been deactivated. Please contact support.')
                return render(request, 'unified_login.html')
            
            login(request, user)
            
            # Merge guest cart ONLY for customers (not owners or staff)
            if not user.groups.filter(name__in=['Owner', 'Staff']).exists():
                merge_session_cart_to_user(request, user)
            
            # Check for next parameter for redirection
            next_url = request.GET.get('next') or request.POST.get('next')
            if next_url:
                return redirect(next_url)
            
            # Determine user type and redirect to appropriate dashboard
            if user.groups.filter(name='Owner').exists():
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                return redirect('smart_market:owner_dashboard')
            elif user.groups.filter(name='Staff').exists():
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                return redirect('smart_market:staff_dashboard')
            elif user.groups.filter(name='Customer').exists():
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                return redirect('smart_market:customer_dashboard')
            else:
                # Default customer access for users without specific groups
                customer_group, created = Group.objects.get_or_create(name='Customer')
                user.groups.add(customer_group)
                messages.success(request, f'Welcome, {user.get_full_name() or user.username}!')
                return redirect('smart_market:customer_dashboard')
        else:
            # Authentication failed
            messages.error(request, 'Invalid email/username or password. Please try again or contact support if you need help.')
            return render(request, 'unified_login.html')
    
    return render(request, 'unified_login.html')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Owner').exists(), login_url='smart_market:unified_login')
def owner_dashboard(request):
    """Enhanced owner dashboard with comprehensive business metrics and ML insights"""
    try:
        # ---- Business Metrics ----
        sales_total = Order.objects.aggregate(total=Sum('total_amount'))['total'] or 0
        orders_count = Order.objects.count()
        pending_orders = Order.objects.filter(status='pending').count()
        completed_orders = Order.objects.filter(status='delivered').count()

        # User counts
        customer_count = User.objects.filter(groups__name='Customer').count()
        staff_count = User.objects.filter(groups__name='Staff').count()
        total_users = customer_count + staff_count

        # Product inventory
        total_products = Product.objects.count() + SmartProducts.objects.count()
        low_stock_products = list(Product.objects.filter(stock_quantity__lt=10)) + \
                           list(SmartProducts.objects.filter(stock_quantity__lt=10))

        # Top products by sales volume (from order items)
        top_products = list(
            OrderItem.objects.values('product_name')
            .annotate(total_qty=Sum('quantity'), total_rev=Sum('subtotal'))
            .order_by('-total_qty')[:5]
        )

        # Recent orders - show more for scrolling
        recent_orders = Order.objects.order_by('-created_at')[:20]
        
        # All orders for scrollable view
        all_orders = Order.objects.order_by('-created_at')[:50]  # Limit to 50 for performance
        
        # Pending orders for the slide filter
        pending_orders_list = Order.objects.filter(status='pending').order_by('-created_at')[:10]

        # Notifications
        notifications = Notification.objects.filter(recipient_user=request.user).order_by('-created_at')[:10]
        unread_notifications_count = get_unread_notifications_count(request.user)

        # ---- ML / Forecast Metrics ----
        ml_models = MLForecastModel.objects.filter(is_active=True).order_by('-accuracy_score')
        best_model = ml_models.first()

        # Products with active predictions
        predicted_smart_count = SmartProducts.objects.filter(predicted_demand_7d__gt=0).count()
        predicted_regular_count = Product.objects.filter(predicted_demand_7d__gt=0).count()
        total_predicted = predicted_smart_count + predicted_regular_count

        avg_forecast_accuracy = SmartProducts.objects.filter(
            forecast_accuracy__gt=0
        ).aggregate(avg=Avg('forecast_accuracy'))['avg'] or 0

        last_forecast_update = SmartProducts.objects.aggregate(
            latest=Max('last_forecast_update')
        )['latest']

        # Top predicted demand products
        top_demand_products = list(
            SmartProducts.objects.filter(predicted_demand_7d__gt=0)
            .order_by('-predicted_demand_7d')[:5]
        )

        # High-demand + low-stock alert items
        restock_alerts_qs = SmartProducts.objects.filter(
            predicted_demand_7d__gt=0,
            stock_quantity__lt=F('predicted_demand_7d')
        ).order_by('-predicted_demand_7d')[:5]
        restock_alerts = []
        for p in restock_alerts_qs:
            p.shortfall = p.predicted_demand_7d - (p.stock_quantity or 0)
            restock_alerts.append(p)

        context = {
            'owner_name': request.user.get_full_name() or request.user.username,
            'user_role': 'Owner',
            # Business cards
            'sales_total': sales_total,
            'customer_count': customer_count,
            'total_users': total_users,
            'orders_count': orders_count,
            'pending_orders': pending_orders,
            'completed_orders': completed_orders,
            'total_products': total_products,
            'low_stock_count': len(low_stock_products),
            'low_stock_products': low_stock_products[:5],
            # Top & recent
            'top_products': top_products,
            'recent_orders': recent_orders,
            'all_orders': all_orders,
            'pending_orders_list': pending_orders_list,
            # Notifications
            'notifications': notifications,
            'unread_notifications_count': unread_notifications_count,
            # ML metrics
            'best_model': best_model,
            'ml_models_count': ml_models.count(),
            'total_predicted': total_predicted,
            'avg_forecast_accuracy': round(float(avg_forecast_accuracy), 1),
            'last_forecast_update': last_forecast_update,
            'top_demand_products': top_demand_products,
            'restock_alerts': restock_alerts,
        }

    except Exception as e:
        context = {
            'owner_name': request.user.get_full_name() or request.user.username,
            'error_message': 'Some dashboard data is temporarily unavailable.',
            'sales_total': 0,
            'orders_count': 0,
            'customer_count': 0,
            'user_role': 'Owner',
        }

    return render(request, 'owner_dashboard.html', context)

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Owner').exists(), login_url='smart_market:unified_login')
def owner_inventory_view(request):
    """Enhanced inventory management view for owners with full CRUD operations"""
    try:
        # Get all products for inventory management
        regular_products = Product.objects.all().order_by('name')
        smart_products = SmartProducts.objects.all().order_by('name')
        
        # Combine all products with type indicator
        all_products = []
        for product in regular_products:
            all_products.append({
                'product': product,
                'type': 'regular',
                'low_stock': (product.stock_quantity or 0) < 10
            })
        for product in smart_products:
            all_products.append({
                'product': product,
                'type': 'smart',
                'low_stock': (product.stock_quantity or 0) < 10
            })
        
        # Filter options
        search_query = request.GET.get('search', '').strip()
        if search_query:
            all_products = [item for item in all_products if search_query.lower() in item['product'].name.lower()]
        
        show_low_stock = request.GET.get('low_stock_only', '') == 'true'
        if show_low_stock:
            all_products = [item for item in all_products if item['low_stock']]
        
        # Statistics
        total_products_count = len(all_products)
        low_stock_count = len([item for item in all_products if item['low_stock']])
        total_inventory_value = sum(float(item['product'].price * (item['product'].stock_quantity or 0)) for item in all_products)
        
        # Get categories for add/edit forms
        categories = Category.objects.all()
        
        context = {
            'owner_name': request.user.get_full_name() or request.user.username,
            'products': all_products,
            'total_products_count': total_products_count,
            'low_stock_count': low_stock_count,
            'total_inventory_value': total_inventory_value,
            'search_query': search_query,
            'show_low_stock': show_low_stock,
            'categories': categories
        }
        
    except Exception as e:
        context = {
            'owner_name': request.user.get_full_name() or request.user.username,
            'error_message': 'Inventory data temporarily unavailable.',
            'products': [],
            'categories': []
        }
    
    return render(request, 'owner_inventory.html', context)

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Owner').exists(), login_url='smart_market:unified_login')
@require_POST
def add_product(request):
    """Add new product to inventory"""
    try:
        name = request.POST.get('name', '').strip()
        price = request.POST.get('price', '')
        stock_quantity = request.POST.get('stock_quantity', '0')
        description = request.POST.get('description', '').strip()
        image_url = request.POST.get('image_url', '').strip()
        product_type = request.POST.get('type', '')
        
        # Validation
        if not name or not price or not product_type:
            messages.error(request, 'Product name, price, and type are required.')
            return redirect('smart_market:owner_inventory')
        
        try:
            price = float(price)
            stock_quantity = int(stock_quantity) if stock_quantity else 0
        except ValueError:
            messages.error(request, 'Invalid price or stock quantity.')
            return redirect('smart_market:owner_inventory')
        
        if product_type == 'regular':
            Product.objects.create(
                name=name,
                description=description,
                price=price,
                stock_quantity=stock_quantity,
                image_url=image_url,
                in_stock=stock_quantity > 0
            )
        else:  # smart product
            SmartProducts.objects.create(
                name=name,
                description=description,
                price=price,
                stock_quantity=stock_quantity,
                image_url=image_url
            )
        
        messages.success(request, f'Product "{name}" added successfully!')
        
    except Exception as e:
        messages.error(request, f'Error adding product: {str(e)}')
    
    return redirect('smart_market:owner_inventory')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Owner').exists(), login_url='smart_market:unified_login')
@require_POST
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Owner').exists(), login_url='smart_market:unified_login')
@require_POST
def edit_product(request):
    """Edit existing product"""
    try:
        product_id = request.POST.get('product_id')
        product_type = request.POST.get('product_type')
        
        if product_type == 'regular':
            product = get_object_or_404(Product, id=product_id)
        else:
            product = get_object_or_404(SmartProducts, id=product_id)
        
        # Update fields
        product.name = request.POST.get('name', product.name).strip()
        product.description = request.POST.get('description', product.description or '')
        
        try:
            product.price = float(request.POST.get('price', product.price))
            product.stock_quantity = int(request.POST.get('stock_quantity', product.stock_quantity or 0))
        except ValueError:
            messages.error(request, 'Invalid price or stock quantity.')
            return redirect('smart_market:owner_inventory')
        
        product.image_url = request.POST.get('image_url', product.image_url or '')
        
        if product_type == 'regular':
            product.in_stock = product.stock_quantity > 0
        
        product.save()
        messages.success(request, f'Product "{product.name}" updated successfully!')
        
    except Exception as e:
        messages.error(request, f'Error updating product: {str(e)}')
    
    return redirect('smart_market:owner_inventory')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Owner').exists(), login_url='smart_market:unified_login')
@require_POST
def delete_product(request):
    """Delete product from inventory"""
    try:
        product_id = request.POST.get('product_id')
        product_type = request.POST.get('product_type')
        
        if product_type == 'regular':
            product = get_object_or_404(Product, id=product_id)
        else:
            product = get_object_or_404(SmartProducts, id=product_id)
        
        product_name = product.name
        product.delete()
        
        messages.success(request, f'Product "{product_name}" deleted successfully!')
        
    except Exception as e:
        messages.error(request, f'Error deleting product: {str(e)}')
    
    return redirect('smart_market:owner_inventory')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Owner').exists(), login_url='smart_market:unified_login')
@require_POST
def update_stock(request):
    """Update product stock quantity"""
    try:
        product_id = request.POST.get('product_id')
        product_type = request.POST.get('product_type')
        new_stock = request.POST.get('stock_quantity')
        
        try:
            new_stock = int(new_stock)
        except ValueError:
            messages.error(request, 'Invalid stock quantity.')
            return redirect('smart_market:owner_inventory')
        
        if product_type == 'regular':
            product = get_object_or_404(Product, id=product_id)
            product.stock_quantity = new_stock
            product.in_stock = new_stock > 0
        else:
            product = get_object_or_404(SmartProducts, id=product_id)
            product.stock_quantity = new_stock
        
        product.save()
        messages.success(request, f'Stock updated for "{product.name}" to {new_stock} items.')
        
    except Exception as e:
        messages.error(request, f'Error updating stock: {str(e)}')
    
    return redirect('smart_market:owner_inventory')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Owner').exists(), login_url='smart_market:unified_login')
def all_orders_view(request):
    """View all customer orders for owner management"""
    try:
        # Get filter parameters
        status_filter = request.GET.get('status', '')
        search_query = request.GET.get('search', '').strip()
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        # Start with all orders
        orders = Order.objects.all()
        
        # Apply filters
        if status_filter:
            orders = orders.filter(status=status_filter)
            
        if search_query:
            orders = orders.filter(
                models.Q(order_number__icontains=search_query) |
                models.Q(customer_name__icontains=search_query) |
                models.Q(customer_email__icontains=search_query)
            )
            
        if date_from:
            orders = orders.filter(created_at__gte=date_from)
            
        if date_to:
            orders = orders.filter(created_at__lte=date_to)
        
        # Order by most recent first
        orders = orders.order_by('-created_at')
        
        # Calculate summary statistics
        total_orders = orders.count()
        total_revenue = sum(order.total_amount for order in orders)
        status_counts = {
            'pending': orders.filter(status='pending').count(),
            'processing': orders.filter(status='processing').count(),
            'shipped': orders.filter(status='shipped').count(),
            'delivered': orders.filter(status='delivered').count(),
            'cancelled': orders.filter(status='cancelled').count(),
        }
        
        context = {
            'orders': orders,
            'total_orders': total_orders,
            'total_revenue': total_revenue,
            'status_counts': status_counts,
            'status_filter': status_filter,
            'search_query': search_query,
            'date_from': date_from,
            'date_to': date_to,
            'order_statuses': Order.STATUS_CHOICES,
        }
        
    except Exception as e:
        messages.error(request, f'Error loading orders: {str(e)}')
        context = {
            'orders': [],
            'total_orders': 0,
            'total_revenue': 0,
            'status_counts': {},
            'order_statuses': Order.STATUS_CHOICES,
        }
    
    return render(request, 'all_orders.html', context)


@login_required
@require_POST
def update_order_status(request, order_number):
    """AJAX endpoint to update order status — owner/staff only."""
    if not (request.user.is_superuser or request.user.groups.filter(name__in=['Owner', 'Staff']).exists()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    order = get_object_or_404(Order, order_number=order_number)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

    new_status = data.get('status', '').strip()
    valid_statuses = [s[0] for s in Order.STATUS_CHOICES]
    if new_status not in valid_statuses:
        return JsonResponse({'success': False, 'error': f'Invalid status: {new_status}'}, status=400)

    ALLOWED_TRANSITIONS = {
        'pending':    ['processing', 'cancelled'],
        'processing': ['shipped', 'cancelled'],
        'shipped':    ['delivered'],
        'delivered':  ['refunded'],
        'cancelled':  [],
        'refunded':   [],
    }
    if new_status not in ALLOWED_TRANSITIONS.get(order.status, []):
        return JsonResponse({
            'success': False,
            'error': f'Cannot change from {order.get_status_display()} to {new_status}'
        }, status=400)

    old_status = order.get_status_display()
    order.status = new_status
    order.save(update_fields=['status', 'updated_at'])

    # Create notification for the customer
    try:
        Notification.objects.create(
            recipient_user=order.user,
            notification_type='order_update',
            title=f'Order {order.order_number} Updated',
            message=f'Your order status changed from {old_status} to {order.get_status_display()}.',
            related_order=order,
        )
    except Exception:
        pass  # Don't fail the status update if notification fails

    return JsonResponse({
        'success': True,
        'new_status': new_status,
        'new_status_display': order.get_status_display(),
        'order_number': order.order_number,
    })


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Staff').exists(), login_url='smart_market:unified_login')
def staff_dashboard(request):
    """Enhanced staff dashboard with relevant information"""
    try:
        # Basic statistics
        total_products = Product.objects.count() + SmartProducts.objects.count()
        customer_count = User.objects.filter(groups__name='Customer').count()
        recent_orders = Order.objects.order_by('-created_at')[:10]
        pending_orders = Order.objects.filter(status='pending').count()
        
        # Low stock alerts
        low_stock_products = list(Product.objects.filter(stock_quantity__lt=10)) + \
                           list(SmartProducts.objects.filter(stock_quantity__lt=10))
        
        context = {
            'staff_name': request.user.get_full_name() or request.user.username,
            'user_role': 'Staff',
            'total_products': total_products,
            'customer_count': customer_count,
            'pending_orders': pending_orders,
            'recent_orders': recent_orders,
            'low_stock_count': len(low_stock_products),
            'low_stock_products': low_stock_products[:10],
            'staff_stats': {
                'alerts': f'{len(low_stock_products)} low stock items',
                'orders_to_process': f'{pending_orders} pending orders',
                'customer_base': f'{customer_count} registered customers'
            }
        }
        
    except Exception as e:
        context = {
            'staff_name': request.user.get_full_name() or request.user.username,
            'user_role': 'Staff',
            'error_message': 'Some dashboard data is temporarily unavailable.'
        }
    
    return render(request, 'staff_dashboard.html', context)

@login_required
def customer_dashboard(request):
    """Enhanced customer dashboard with personalized information"""
    # Ensure only customers or users without owner/staff privileges can access
    if request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        messages.warning(request, 'Access denied. Please login with customer credentials.')
        return redirect('smart_market:unified_login')
    
    try:
        # Get or create customer cart
        cart, created = Cart.objects.get_or_create(user=request.user)
        
        # Get customer's orders
        customer_orders = Order.objects.filter(user=request.user).order_by('-created_at')
        recent_orders = customer_orders[:5]
        
        # Calculate customer statistics
        total_orders = customer_orders.count()
        total_spent = sum(order.total_amount for order in customer_orders if order.total_amount)
        pending_orders = customer_orders.filter(status='pending').count()
        
        # Calculate average order value
        average_order_value = (total_spent / total_orders) if total_orders > 0 else 0
        
        # Get cart information
        cart_items = cart.items.all() if hasattr(cart, 'items') else []
        cart_total = cart.total_amount if hasattr(cart, 'total_amount') else 0
        cart_item_count = cart.total_items if hasattr(cart, 'total_items') else len(cart_items)
        
        # Recent products for recommendations
        available_products = list(Product.objects.filter(in_stock=True)[:6]) + \
                           list(SmartProducts.objects.exclude(stock_quantity=0)[:6])

        context = {
            'customer_name': request.user.get_full_name() or request.user.username,
            'customer_email': request.user.email,
            'user_role': 'Customer',
            'total_orders': total_orders,
            'total_spent': total_spent,
            'average_order_value': average_order_value,
            'pending_orders': pending_orders,
            'recent_orders': recent_orders,
            'cart_total': cart_total,
            'cart_item_count': cart_item_count,
            'cart_items': cart_items,
            'recommended_products': available_products,
            'customer_stats': {
                'member_since': request.user.date_joined.strftime('%B %Y'),
                'order_status': f'{total_orders} orders placed',
                'loyalty_status': 'Premium' if total_orders > 5 else 'Standard'
            }
        }
        
    except Exception as e:
        # Fallback data
        context = {
            'customer_name': request.user.get_full_name() or request.user.username,
            'customer_email': request.user.email,
            'user_role': 'Customer',
            'error_message': 'Some dashboard data is temporarily unavailable.',
            'total_orders': 0,
            'total_spent': 0,
            'average_order_value': 0,
            'cart_total': 0,
            'cart_item_count': 0
        }
    
    return render(request, 'customer_dashboard.html', context)

def logout_view(request):
    if request.method == 'POST':
        # Process the logout
        if request.user.is_authenticated:
            # Get user's display name (first name if available, otherwise username)
            display_name = request.user.first_name if request.user.first_name else request.user.username
        else:
            display_name = "User"
        logout(request)
        
        # Render logout confirmation page
        context = {
            'display_name': display_name,
            'logged_out': True
        }
        return render(request, 'logout_confirmation.html', context)
    
    elif request.method == 'GET':
        if request.user.is_authenticated:
            # Show logout confirmation form
            # Get user's display name (first name if available, otherwise username)
            display_name = request.user.first_name if request.user.first_name else request.user.username
            context = {
                'display_name': display_name,
                'logged_out': False
            }
            return render(request, 'logout_confirmation.html', context)
        else:
            # User not logged in, redirect to login
            return redirect('smart_market:login')
def customer_signup(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        if not name or not email or not password:
            return render(request, 'customer_signup.html', {'error': 'All fields are required.'})
        if User.objects.filter(email=email).exists():
            return render(request, 'customer_signup.html', {'error': 'Email already registered.'})
        if User.objects.filter(username=email).exists():
            return render(request, 'customer_signup.html', {'error': 'Username already taken.'})
        user = User.objects.create_user(username=email, email=email, password=password, first_name=name)
        # Add user to Customer group
        customer_group, created = Group.objects.get_or_create(name='Customer')
        user.groups.add(customer_group)
        user.save()
        login(request, user)
        
        # Merge any guest cart items to the new user account
        merge_session_cart_to_user(request, user)
        
        # Check for next parameter for redirection
        next_url = request.GET.get('next') or request.POST.get('next')
        if next_url:
            return redirect(next_url)
        
        return redirect('smart_market:customer_dashboard')
    return render(request, 'customer_signup.html')

from django.shortcuts import render, get_object_or_404, redirect
from .models import Product, Category, SmartProducts
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail
from django.urls import reverse
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from django.conf import settings
from django.http import JsonResponse


# Seasonal Product Recommendation API endpoint
def get_recommendations_api(request):
    """API endpoint for fetching ML-powered recommendations"""
    recommendation_type = request.GET.get('type', 'seasonal')
    limit = min(int(request.GET.get('limit', 6)), 20)  # Max 20 products
    
    if recommendation_type == 'seasonal':
        products = get_seasonal_recommendations(limit=limit)
    elif recommendation_type == 'weather':
        products = get_weather_based_recommendations()[:limit]
    elif recommendation_type == 'trending':
        products = SmartProducts.objects.filter(
            predicted_demand_7d__gt=40
        ).order_by('-predicted_demand_7d')[:limit]
        
        for product in products:
            product.recommendation_reason = "High demand this week"
            product.recommendation_type = "trending"
    elif recommendation_type == 'discount':
        products = SmartProducts.objects.filter(
            is_promotional=True
        ).order_by('-promotion_lift')[:limit]
        
        for product in products:
            discount_pct = (float(product.promotion_lift) - 1.0) * 100 if product.promotion_lift > 1 else 0
            product.recommendation_reason = f"Special promotion - {discount_pct:.0f}% off"
            product.recommendation_type = "discount"
    else:
        products = []
    
    # Convert to JSON response
    products_data = []
    for product in products:
        product_data = {
            'id': product.id,
            'name': product.name,
            'price': str(product.price),
            'image_url': product.image_url if hasattr(product, 'image_url') else '',
            'category': product.category if hasattr(product, 'category') else '',
            'recommendation_reason': getattr(product, 'recommendation_reason', ''),
            'recommendation_type': getattr(product, 'recommendation_type', recommendation_type),
            'predicted_demand_7d': getattr(product, 'predicted_demand_7d', None),
            'seasonal_coefficient': getattr(product, 'seasonal_coefficient', None),
        }
        products_data.append(product_data)
    
    return JsonResponse({
        'products': products_data,
        'type': recommendation_type,
        'current_season': get_current_season(),
        'upcoming_festivals': get_upcoming_festivals(),
        'is_weekend': datetime.now().weekday() >= 5
    })





def home(request):
    q = request.GET.get('q', '').strip()
    user = request.user if request.user.is_authenticated else None
    
    if q:
        # Search in both Product and SmartProducts models
        regular_products = Product.objects.filter(name__icontains=q)
        smart_products = SmartProducts.objects.filter(name__icontains=q)
        
        # Combine the products
        all_products = list(regular_products) + list(smart_products)
        products = all_products[:12]  # Limit to 12 total products
        
        # No recommendations during search
        seasonal_recommendations = []
        trending_products = []
        weather_recommendations = []
        personal_recommendations = []
    else:
        # Get regular products for display
        regular_products = Product.objects.all()[:3]
        smart_products = SmartProducts.objects.all()[:6]
        products = list(regular_products) + list(smart_products)
        products = products[:9]
        
        # Get ML-powered recommendations (passes user for personalization)
        seasonal_recommendations = get_seasonal_recommendations(limit=6, user=user)
        
        # Get trending products — ordered by REAL demand from orders
        trending_products = SmartProducts.objects.filter(
            predicted_demand_7d__gt=0
        ).order_by('-predicted_demand_7d')[:4]
        
        for product in trending_products:
            product.recommendation_reason = f"Trending — {product.predicted_demand_7d} sold this week"
            product.recommendation_type = "trending"
        
        # Get weather-based recommendations
        weather_recommendations = get_weather_based_recommendations(user=user)
        
        # Get user-personalized recommendations
        personal_recommendations = get_user_personal_recommendations(user, limit=4)
    
    # Get current context information
    current_season = get_current_season()
    upcoming_festivals = get_upcoming_festivals()
    is_weekend = datetime.now().weekday() >= 5
    
    context = {
        'products': products,
        'q': q,
        'seasonal_recommendations': seasonal_recommendations,
        'trending_products': trending_products,
        'weather_recommendations': weather_recommendations,
        'personal_recommendations': personal_recommendations,
        'current_season': current_season,
        'upcoming_festivals': upcoming_festivals,
        'is_weekend': is_weekend,
        'season_display': current_season.title() if current_season != 'all_year' else 'All Seasons'
    }
    
    return render(request, 'home.html', context)

def contact(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        question = request.POST.get('question', '').strip()
        
        if not username or not email or not question:
            return render(request, 'contact.html', {'error': 'All fields are required.'})
        
        ticket_number = f'TKT-{random.randint(100000, 999999)}'

        # Notify all owners about the new support ticket
        create_notification(
            notification_type='support_ticket',
            title=f'New Support Ticket {ticket_number}',
            message=f'From: {username} ({email})\nMessage: {question[:200]}',
            recipient_groups=['Owner'],
        )

        return render(request, 'contact.html', {
            'success': True,
            'ticket_number': ticket_number,
            'submitted_name': username,
            'submitted_email': email,
        })
    
    return render(request, 'contact.html')

def whatsapp(request):
    user_message = None
    if request.method == 'POST':
        user_message = request.POST.get('message', '').strip()
    return render(request, 'whatsapp.html', {'user_message': user_message})

    context = {
        'products': products,
        'query': q,
    }
    return render(request, 'home.html', context)


def shop(request):
    """Product catalog with shopping cart functionality"""
    # Get all products from both models
    regular_products = Product.objects.all()
    smart_products = SmartProducts.objects.all()
    
    # Combine all products
    all_products = list(regular_products) + list(smart_products)
    
    # Get categories from both models
    regular_categories = Product.objects.filter(category__isnull=False).values_list('category__name', flat=True).distinct()
    smart_categories = SmartProducts.objects.exclude(category__isnull=True).exclude(category='').values_list('category', flat=True).distinct()
    
    # Combine and create categories list
    all_categories = list(regular_categories) + list(smart_categories)
    categories = [{'name': cat, 'id': cat} for cat in set(all_categories) if cat and cat.strip()]
    
    # Sort categories alphabetically
    categories = sorted(categories, key=lambda x: x['name'])
    
    # Apply search filter
    search_query = request.GET.get('search', '').strip()
    if search_query:
        filtered_products = []
        for product in all_products:
            if (search_query.lower() in product.name.lower() or 
                (hasattr(product, 'description') and product.description and 
                 search_query.lower() in product.description.lower())):
                filtered_products.append(product)
        all_products = filtered_products
    
    # Apply category filter
    category_filter = request.GET.get('category', '')
    if category_filter:
        filtered_products = []
        for product in all_products:
            # Handle Product model (ForeignKey to Category)
            if hasattr(product, 'category') and hasattr(product.category, 'name'):
                product_category = product.category.name
            # Handle SmartProducts model (string category) 
            elif hasattr(product, 'category'):
                product_category = product.category
            else:
                product_category = None
                
            if product_category == category_filter:
                filtered_products.append(product)
        all_products = filtered_products
    
    # Price range filter
    min_price = request.GET.get('min_price', '')
    max_price = request.GET.get('max_price', '')
    
    if min_price or max_price:
        filtered_products = []
        for product in all_products:
            product_price = float(product.price) if product.price else 0
            price_match = True
            
            if min_price:
                try:
                    if product_price < float(min_price):
                        price_match = False
                except (ValueError, TypeError):
                    pass
                    
            if max_price and price_match:
                try:
                    if product_price > float(max_price):
                        price_match = False
                except (ValueError, TypeError):
                    pass
                    
            if price_match:
                filtered_products.append(product)
        all_products = filtered_products
    
    # Get maximum price for the price filter
    max_price_available = 0
    for product in Product.objects.all():
        if product.price and float(product.price) > max_price_available:
            max_price_available = float(product.price)
    for product in SmartProducts.objects.all():
        if product.price and float(product.price) > max_price_available:
            max_price_available = float(product.price)
    
    # Add ML-powered recommendations - only show when not actively filtering
    show_recommendations = not (search_query or category_filter or min_price or max_price)
    
    if show_recommendations:
        user = request.user if request.user.is_authenticated else None
        
        # Get seasonal recommendations (ML-powered with fallback)
        seasonal_recommendations = get_seasonal_recommendations(limit=4, user=user)
        
        # Get trending products — ordered by REAL demand from orders
        trending_products = SmartProducts.objects.filter(
            predicted_demand_7d__gt=0
        ).order_by('-predicted_demand_7d')[:4]
        
        for product in trending_products:
            product.recommendation_reason = f"Trending — {product.predicted_demand_7d} sold this week"
            product.recommendation_type = "trending"
        
        # Get weather-based recommendations (ML-powered with fallback)
        weather_recommendations = get_weather_based_recommendations(user=user)[:3]
        
        # Get products on sale or recently restocked
        discounted_products = SmartProducts.objects.filter(
            is_promotional=True
        ).order_by('-promotion_lift')[:3]
        
        for product in discounted_products:
            discount_pct = (float(product.promotion_lift) - 1.0) * 100 if product.promotion_lift > 1 else 0
            product.recommendation_reason = f"Special promotion - {discount_pct:.0f}% off"
            product.recommendation_type = "discount"
        
        # User-personalized recommendations
        personal_recommendations = get_user_personal_recommendations(user, limit=4)
    else:
        seasonal_recommendations = []
        trending_products = []
        weather_recommendations = []
        discounted_products = []
        personal_recommendations = []
    
    # Get current context information
    current_season = get_current_season()
    upcoming_festivals = get_upcoming_festivals()
    is_weekend = datetime.now().weekday() >= 5
    
    context = {
        'products': all_products,
        'categories': categories,
        'search_query': search_query,
        'selected_category': category_filter,
        'min_price': min_price,
        'max_price': max_price,
        'max_price_available': max_price_available,
        'seasonal_recommendations': seasonal_recommendations,
        'trending_products': trending_products,
        'weather_recommendations': weather_recommendations,
        'discounted_products': discounted_products,
        'personal_recommendations': personal_recommendations,
        'show_recommendations': show_recommendations,
        'current_season': current_season,
        'upcoming_festivals': upcoming_festivals,
        'is_weekend': is_weekend,
        'season_display': current_season.title() if current_season != 'all_year' else 'All Seasons'
    }
    return render(request, 'shop.html', context)


def search(request):
    # simple search endpoint that reuses home template
    return home(request)


# Export ALL business records to Excel for owner
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def _style_header_row(ws, num_cols):
    """Apply header styling to the first row of a worksheet."""
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    thin_border = Border(
        bottom=Side(style='thin', color='000000'),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

def _auto_width(ws):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 50)

from openpyxl.worksheet.table import Table, TableStyleInfo

def _make_table(ws, table_name):
    """Convert sheet data into a proper Excel Table so Power BI can load it."""
    if ws.max_row < 2:
        ws.append([None] * ws.max_column)
    max_col_letter = openpyxl.utils.get_column_letter(ws.max_column)
    ref = f"A1:{max_col_letter}{ws.max_row}"
    safe_name = table_name.replace(' ', '_').replace('-', '_')
    tab = Table(displayName=safe_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)

def _safe(val):
    """Convert value to Excel-safe type. None stays None for Power BI."""
    if val is None:
        return None
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(val, dict) or isinstance(val, list):
        return str(val)
    return val

def _text(val):
    """Convert nullable text field: None becomes 'N/A' for consistent Power BI text type."""
    if val is None or val == '':
        return 'N/A'
    return str(val)

def _calculate_real_time_sales_data(product_name):
    """Calculate real-time sales data for a product."""
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    now = timezone.now()
    thirty_days_ago = now - timedelta(days=30)
    seven_days_ago = now - timedelta(days=7)
    
    # Calculate weekly sales (last 7 days)
    weekly_sales = OrderItem.objects.filter(
        product_name=product_name,
        order__created_at__gte=seven_days_ago,
        order__status__in=['pending', 'processing', 'shipped', 'delivered']
    ).aggregate(
        total_qty=models.Sum('quantity'),
        total_revenue=models.Sum('subtotal')
    )
    
    # Calculate monthly sales (last 30 days)
    monthly_sales = OrderItem.objects.filter(
        product_name=product_name,
        order__created_at__gte=thirty_days_ago,
        order__status__in=['pending', 'processing', 'shipped', 'delivered']
    ).aggregate(
        total_qty=models.Sum('quantity'),
        total_revenue=models.Sum('subtotal')
    )
    
    # Calculate averages
    avg_weekly_qty = weekly_sales['total_qty'] or 0
    avg_weekly_revenue = float(weekly_sales['total_revenue'] or 0)
    
    # Calculate monthly averages (4-week period)
    avg_monthly_qty = monthly_sales['total_qty'] or 0
    avg_monthly_revenue = float(monthly_sales['total_revenue'] or 0)
    
    return {
        'avg_weekly_qty': avg_weekly_qty,
        'avg_weekly_revenue': avg_weekly_revenue,
        'avg_monthly_qty': avg_monthly_qty,
        'avg_monthly_revenue': avg_monthly_revenue,
        'last_calculated': now.strftime('%Y-%m-%d %H:%M:%S')
    }

def _get_stock_status_info(product):
    """Get comprehensive stock status information."""
    stock_qty = product.stock_quantity or 0
    reorder_point = product.reorder_point or 0
    min_stock = product.min_stock_level or 0
    max_stock = product.max_stock_level or 0
    
    status = 'Normal'
    if stock_qty <= 0:
        status = 'Out of Stock'
    elif stock_qty <= min_stock:
        status = 'Critical Low'
    elif stock_qty <= reorder_point:
        status = 'Needs Reorder'
    elif stock_qty > max_stock * 0.9:
        status = 'Overstocked'
    
    days_of_stock = 0
    if hasattr(product, 'avg_weekly_sales') and product.avg_weekly_sales:
        daily_sales = float(product.avg_weekly_sales) / 7
        if daily_sales > 0:
            days_of_stock = round(stock_qty / daily_sales, 1)
    
    return {
        'status': status,
        'days_of_stock': days_of_stock,
        'stock_coverage': 'N/A' if days_of_stock == 0 else f"{days_of_stock} days"
    }

@login_required
def export_dashboard_excel(request):
    """Export ALL business records to Excel with real-time data — owner only."""
    if not (request.user.is_superuser or request.user.groups.filter(name='Owner').exists()):
        messages.error(request, "Only the business owner can export records.")
        return redirect('smart_market:home')

    from .models import (
        Product, SmartProducts, Category, Order, OrderItem, Payment,
        Customers, Employees, Suppliers, Inventory, CustomerSupport,
        DailySales, StoreInfo, AuditLog, Notification, Reviews,
        SeasonalSalesData, MLForecastModel, ForecastPrediction,
        ProductRecommendationLog,
    )

    wb = openpyxl.Workbook()

    # ── 1. Dashboard Summary ──────────────────────────────
    ws = wb.active
    ws.title = 'Dashboard Summary'
    orders = Order.objects.all()
    total_revenue = orders.aggregate(s=Sum('total_amount'))['s'] or 0
    total_orders = orders.count()
    total_products = Product.objects.count() + SmartProducts.objects.count()
    total_customers = User.objects.filter(groups__name='Customer').count()
    headers = ['Metric', 'Value']
    ws.append(headers)
    # Calculate real-time inventory alerts
    products_low_stock = Product.objects.filter(
        stock_quantity__isnull=False, 
        stock_quantity__lte=models.F('reorder_point')
    ).count()
    smart_products_low_stock = SmartProducts.objects.filter(
        stock_quantity__isnull=False, 
        stock_quantity__lte=models.F('reorder_point')
    ).count()
    
    products_out_of_stock = Product.objects.filter(
        stock_quantity__isnull=False,
        stock_quantity__lte=0
    ).count()
    smart_products_out_of_stock = SmartProducts.objects.filter(
        stock_quantity__isnull=False,
        stock_quantity__lte=0
    ).count()
    
    # Calculate forecast data freshness
    outdated_forecasts = Product.objects.filter(
        last_forecast_update__isnull=True
    ).count() + SmartProducts.objects.filter(
        last_forecast_update__isnull=True
    ).count()
    
    summary_rows = [
        ('Total Revenue (Rs)', str(_safe(total_revenue))),
        ('Total Orders', str(total_orders)),
        ('Total Products (Regular)', str(Product.objects.count())),
        ('Total Products (Smart)', str(SmartProducts.objects.count())),
        ('Total Customers', str(total_customers)),
        ('Total Employees', str(Employees.objects.count())),
        ('Total Suppliers', str(Suppliers.objects.count())),
        ('Pending Orders', str(orders.filter(status='pending').count())),
        ('Processing Orders', str(orders.filter(status='processing').count())),
        ('Delivered Orders', str(orders.filter(status='delivered').count())),
        ('Cancelled Orders', str(orders.filter(status='cancelled').count())),
        ('Open Support Tickets', str(CustomerSupport.objects.filter(status='open').count())),
        ('Products Need Reorder', str(products_low_stock)),
        ('Smart Products Need Reorder', str(smart_products_low_stock)),
        ('Products Out of Stock', str(products_out_of_stock)),
        ('Smart Products Out of Stock', str(smart_products_out_of_stock)),
        ('Products with Outdated Forecasts', str(outdated_forecasts)),
        ('Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Data Calculated in Real-Time', 'Yes'),
    ]
    for row in summary_rows:
        ws.append(list(row))
    _style_header_row(ws, 2)
    _auto_width(ws)
    _make_table(ws, 'DashboardSummary')

    # ── 2. Products (with Real-Time Data) ──────────────────────────────────────
    ws2 = wb.create_sheet('Products')
    prod_headers = ['ID', 'Name', 'Category', 'Price', 'Stock Qty', 'Stock Status', 'Stock Coverage',
                    'In Stock', 'Promotional', 'Peak Season', 'Festival', 'Demand Trend',
                    'Current Weekly Sales', 'Current Monthly Sales', 'DB Weekly Sales', 'DB Monthly Sales',
                    'Predicted Demand 7d', 'Predicted Demand 30d', 'Forecast Accuracy', 'Last Forecast Update',
                    'Reorder Point', 'Min Stock', 'Max Stock', 'Seasonal Multiplier', 'Data Updated', 'Created']
    ws2.append(prod_headers)
    
    # Process products with real-time calculations
    products_with_data = []
    for p in Product.objects.select_related('category').all():
        # Get real-time sales data
        sales_data = _calculate_real_time_sales_data(p.name)
        stock_info = _get_stock_status_info(p)
        
        # Format last forecast update
        last_forecast = 'Never' if not p.last_forecast_update else p.last_forecast_update.strftime('%Y-%m-%d %H:%M')
        
        row_data = [
            p.id, p.name, _text(p.category.name if p.category else None),
            _safe(p.price) or 0, p.stock_quantity or 0, stock_info['status'], stock_info['stock_coverage'],
            'Yes' if p.in_stock else 'No', 'Yes' if p.is_promotional else 'No', 
            _text(p.peak_season), _text(p.festival_association), _text(p.demand_trend),
            sales_data['avg_weekly_qty'], sales_data['avg_monthly_qty'],
            _safe(p.avg_weekly_sales) or 0, _safe(p.avg_monthly_sales) or 0,
            p.predicted_demand_7d or 0, p.predicted_demand_30d or 0, _safe(p.forecast_accuracy) or 0,
            last_forecast, p.reorder_point or 0, p.min_stock_level or 0, p.max_stock_level or 0,
            p.get_current_season_multiplier(), sales_data['last_calculated'],
            _safe(p.created_at) or 'N/A'
        ]
        products_with_data.append(row_data)
    
    # Sort by stock status priority (Critical first)
    status_priority = {'Out of Stock': 1, 'Critical Low': 2, 'Needs Reorder': 3, 'Normal': 4, 'Overstocked': 5}
    products_with_data.sort(key=lambda x: status_priority.get(x[5], 6))
    
    for row_data in products_with_data:
        ws2.append(row_data)
    _style_header_row(ws2, len(prod_headers))
    _auto_width(ws2)
    _make_table(ws2, 'Products')

    # ── 3. Smart Products (with Real-Time Data) ────────────────────────
    ws3 = wb.create_sheet('Smart Products')
    sp_headers = ['ID', 'Name', 'Category', 'Price', 'Stock Qty', 'Stock Status', 'Stock Coverage',
                  'Promotional', 'Peak Season', 'Festival', 'Demand Trend',
                  'Current Weekly Sales', 'Current Monthly Sales', 'DB Weekly Sales', 'DB Monthly Sales',
                  'Predicted Demand 7d', 'Predicted Demand 30d', 'Forecast Accuracy', 'Last Forecast Update',
                  'Reorder Point', 'Min Stock', 'Max Stock', 'Seasonal Multiplier', 'Data Updated', 'Created']
    ws3.append(sp_headers)
    
    # Process smart products with real-time calculations
    smart_products_with_data = []
    for p in SmartProducts.objects.all():
        # Get real-time sales data
        sales_data = _calculate_real_time_sales_data(p.name)
        stock_info = _get_stock_status_info(p)
        
        # Format last forecast update
        last_forecast = 'Never' if not p.last_forecast_update else p.last_forecast_update.strftime('%Y-%m-%d %H:%M')
        
        row_data = [
            p.id, p.name, _text(p.category), _safe(p.price) or 0, 
            p.stock_quantity or 0, stock_info['status'], stock_info['stock_coverage'],
            'Yes' if p.is_promotional else 'No', _text(p.peak_season), _text(p.festival_association),
            _text(p.demand_trend), sales_data['avg_weekly_qty'], sales_data['avg_monthly_qty'],
            _safe(p.avg_weekly_sales) or 0, _safe(p.avg_monthly_sales) or 0,
            p.predicted_demand_7d or 0, p.predicted_demand_30d or 0, _safe(p.forecast_accuracy) or 0,
            last_forecast, p.reorder_point or 0, p.min_stock_level or 0, p.max_stock_level or 0,
            p.get_current_season_multiplier(), sales_data['last_calculated'],
            _safe(p.created_at) or 'N/A'
        ]
        smart_products_with_data.append(row_data)
    
    # Sort by stock status priority (Critical first)
    status_priority = {'Out of Stock': 1, 'Critical Low': 2, 'Needs Reorder': 3, 'Normal': 4, 'Overstocked': 5}
    smart_products_with_data.sort(key=lambda x: status_priority.get(x[5], 6))
    
    for row_data in smart_products_with_data:
        ws3.append(row_data)
    _style_header_row(ws3, len(sp_headers))
    _auto_width(ws3)
    _make_table(ws3, 'SmartProducts')

    # ── 4. Categories ─────────────────────────────────────
    ws4 = wb.create_sheet('Categories')
    cat_headers = ['ID', 'Name', 'Description', 'Product Count']
    ws4.append(cat_headers)
    for c in Category.objects.annotate(product_count=Count('products')):
        ws4.append([c.id, c.name, _text(c.description), c.product_count])
    _style_header_row(ws4, len(cat_headers))
    _auto_width(ws4)
    _make_table(ws4, 'Categories')

    # ── 5. Orders ─────────────────────────────────────────
    ws5 = wb.create_sheet('Orders')
    ord_headers = ['Order #', 'Customer', 'Email', 'Phone', 'Status',
                   'Subtotal', 'Tax', 'Shipping', 'Total', 'Shipping Address',
                   'City', 'Postal Code', 'Created', 'Updated']
    ws5.append(ord_headers)
    for o in Order.objects.select_related('user').all():
        ws5.append([
            o.order_number, _text(o.customer_name), _text(o.customer_email), _text(o.customer_phone),
            _text(o.status), _safe(o.subtotal) or 0, _safe(o.tax_amount) or 0, _safe(o.shipping_cost) or 0,
            _safe(o.total_amount) or 0, _text(o.shipping_address), _text(o.shipping_city),
            _text(o.shipping_postal_code), _safe(o.created_at) or 'N/A', _safe(o.updated_at) or 'N/A',
        ])
    _style_header_row(ws5, len(ord_headers))
    _auto_width(ws5)
    _make_table(ws5, 'Orders')

    # ── 6. Order Items ────────────────────────────────────
    ws6 = wb.create_sheet('Order Items')
    oi_headers = ['Order #', 'Product Name', 'Unit Price', 'Quantity', 'Subtotal']
    ws6.append(oi_headers)
    for oi in OrderItem.objects.select_related('order').all():
        ws6.append([
            oi.order.order_number, _text(oi.product_name), _safe(oi.unit_price) or 0,
            oi.quantity or 0, _safe(oi.subtotal) or 0,
        ])
    _style_header_row(ws6, len(oi_headers))
    _auto_width(ws6)
    _make_table(ws6, 'OrderItems')

    # ── 7. Payments ───────────────────────────────────────
    ws7 = wb.create_sheet('Payments')
    pay_headers = ['ID', 'Order #', 'Method', 'Amount', 'Status', 'Transaction ID',
                   'Gateway', 'Card Last 4', 'Card Brand', 'Billing Name',
                   'Created', 'Processed']
    ws7.append(pay_headers)
    for pay in Payment.objects.select_related('order').all():
        ws7.append([
            pay.id, pay.order.order_number, _text(pay.payment_method), _safe(pay.amount) or 0,
            _text(pay.status), _text(pay.transaction_id), _text(pay.payment_gateway),
            _text(pay.card_last_four), _text(pay.card_brand), _text(pay.billing_name),
            _safe(pay.created_at) or 'N/A', _safe(pay.processed_at) if pay.processed_at else 'N/A',
        ])
    _style_header_row(ws7, len(pay_headers))
    _auto_width(ws7)
    _make_table(ws7, 'Payments')

    # ── 8. Customers ──────────────────────────────────────
    ws8 = wb.create_sheet('Customers')
    cust_headers = ['ID', 'Name', 'Email', 'Phone', 'Address', 'Credit Record', 'Created']
    ws8.append(cust_headers)
    for c in Customers.objects.all():
        ws8.append([
            c.id, _text(c.name), _text(c.email), _text(c.phone), _text(c.address),
            _safe(c.credit_record) or 0, _safe(c.created_at) or 'N/A',
        ])
    _style_header_row(ws8, len(cust_headers))
    _auto_width(ws8)
    _make_table(ws8, 'Customers')

    # ── 9. Registered Users ───────────────────────────────
    ws9 = wb.create_sheet('Registered Users')
    user_headers = ['ID', 'Username', 'Email', 'First Name', 'Last Name',
                    'Group', 'Active', 'Date Joined', 'Last Login']
    ws9.append(user_headers)
    for u in User.objects.prefetch_related('groups').all():
        groups = ', '.join(g.name for g in u.groups.all()) or 'N/A'
        ws9.append([
            u.id, _text(u.username), _text(u.email), _text(u.first_name), _text(u.last_name),
            groups, 'Yes' if u.is_active else 'No',
            _safe(u.date_joined) or 'N/A', _safe(u.last_login) or 'N/A',
        ])
    _style_header_row(ws9, len(user_headers))
    _auto_width(ws9)
    _make_table(ws9, 'RegisteredUsers')

    # ── 10. Employees ─────────────────────────────────────
    ws10 = wb.create_sheet('Employees')
    emp_headers = ['ID', 'Name', 'Role', 'Contact', 'Email', 'Hire Date']
    ws10.append(emp_headers)
    for e in Employees.objects.all():
        ws10.append([e.id, _text(e.name), _text(e.role), _text(e.contact_number), _text(e.email),
                     _safe(e.hire_date) or 'N/A'])
    _style_header_row(ws10, len(emp_headers))
    _auto_width(ws10)
    _make_table(ws10, 'Employees')

    # ── 11. Suppliers ─────────────────────────────────────
    ws11 = wb.create_sheet('Suppliers')
    sup_headers = ['ID', 'Name', 'Contact', 'Email', 'Address', 'Created']
    ws11.append(sup_headers)
    for s in Suppliers.objects.all():
        ws11.append([s.id, _text(s.name), _text(s.contact_number), _text(s.email),
                     _text(s.address), _safe(s.created_at) or 'N/A'])
    _style_header_row(ws11, len(sup_headers))
    _auto_width(ws11)
    _make_table(ws11, 'Suppliers')

    # ── 12. Inventory Changes ─────────────────────────────
    ws12 = wb.create_sheet('Inventory Changes')
    inv_headers = ['ID', 'Product', 'Change Type', 'Qty Change', 'Date', 'Supplier', 'Notes']
    ws12.append(inv_headers)
    for i in Inventory.objects.select_related('product', 'supplier').all():
        ws12.append([
            i.id, _text(i.product.name), _text(i.change_type), i.quantity_change or 0,
            _safe(i.change_date) or 'N/A', _text(i.supplier.name if i.supplier else None),
            _text(i.notes),
        ])
    _style_header_row(ws12, len(inv_headers))
    _auto_width(ws12)
    _make_table(ws12, 'InventoryChanges')

    # ── 13. Daily Sales ───────────────────────────────────
    ws13 = wb.create_sheet('Daily Sales')
    ds_headers = ['Date', 'Total Sales (Rs)']
    ws13.append(ds_headers)
    for d in DailySales.objects.order_by('-sales_date'):
        ws13.append([_safe(d.sales_date) or 'N/A', _safe(d.total_sales) or 0])
    _style_header_row(ws13, len(ds_headers))
    _auto_width(ws13)
    _make_table(ws13, 'DailySales')

    # ── 14. Reviews ───────────────────────────────────────
    ws14 = wb.create_sheet('Reviews')
    rev_headers = ['ID', 'Customer', 'Product', 'Rating', 'Comment', 'Date']
    ws14.append(rev_headers)
    for r in Reviews.objects.select_related('customer', 'product').all():
        ws14.append([
            r.id, _text(r.customer.name), _text(r.product.name), r.rating or 0,
            _text(r.comment), _safe(r.created_at) or 'N/A',
        ])
    _style_header_row(ws14, len(rev_headers))
    _auto_width(ws14)
    _make_table(ws14, 'Reviews')

    # ── 15. Customer Support Tickets ──────────────────────
    ws15 = wb.create_sheet('Support Tickets')
    cs_headers = ['ID', 'Customer', 'Subject', 'Description', 'Status',
                  'Created', 'Resolved', 'Handled By']
    ws15.append(cs_headers)
    for t in CustomerSupport.objects.select_related('customer', 'handled_by').all():
        ws15.append([
            t.id, _text(t.customer.name), _text(t.subject), _text(t.description), _text(t.status),
            _safe(t.created_at) or 'N/A', _safe(t.resolved_at) if t.resolved_at else 'N/A',
            _text(t.handled_by.name if t.handled_by else None),
        ])
    _style_header_row(ws15, len(cs_headers))
    _auto_width(ws15)
    _make_table(ws15, 'SupportTickets')

    # ── 16. Notifications ─────────────────────────────────
    ws16 = wb.create_sheet('Notifications')
    notif_headers = ['ID', 'Recipient', 'Type', 'Title', 'Message', 'Read', 'Created']
    ws16.append(notif_headers)
    for n in Notification.objects.select_related('recipient_user').all():
        ws16.append([
            n.id, _text(n.recipient_user.username), _text(n.notification_type), _text(n.title),
            _text(n.message), 'Yes' if n.is_read else 'No', _safe(n.created_at) or 'N/A',
        ])
    _style_header_row(ws16, len(notif_headers))
    _auto_width(ws16)
    _make_table(ws16, 'Notifications')

    # ── 17. Audit Log ─────────────────────────────────────
    ws17 = wb.create_sheet('Audit Log')
    audit_headers = ['ID', 'Table', 'Record ID', 'Old Value', 'New Value', 'Updated By', 'Date']
    ws17.append(audit_headers)
    for a in AuditLog.objects.all():
        ws17.append([
            a.id, _text(a.table_name), a.record_id or 0, _text(a.old_value), _text(a.new_value),
            _text(a.updated_by), _safe(a.updated_at) or 'N/A',
        ])
    _style_header_row(ws17, len(audit_headers))
    _auto_width(ws17)
    _make_table(ws17, 'AuditLog')

    # ── 18. Store Info ────────────────────────────────────
    ws18 = wb.create_sheet('Store Info')
    store_headers = ['ID', 'Store Name', 'Address', 'Contact', 'Email', 'Opening Hours']
    ws18.append(store_headers)
    for s in StoreInfo.objects.all():
        ws18.append([s.id, _text(s.store_name), _text(s.address), _text(s.contact_number),
                     _text(s.email), _text(s.opening_hours)])
    _style_header_row(ws18, len(store_headers))
    _auto_width(ws18)
    _make_table(ws18, 'StoreInfo')

    # ══════════════════════════════════════════════════════
    # POWER BI ANALYTICS SHEETS
    # ══════════════════════════════════════════════════════
    from datetime import timedelta
    from django.utils import timezone
    pbi_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')

    def _style_pbi_header(sheet, cols):
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for col in range(1, cols + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = pbi_fill
            cell.alignment = Alignment(horizontal='center')

    # ── PBI-1. Sales Time Series (daily) ──────────────────
    ws_ts = wb.create_sheet('PBI Sales TimeSeries')
    ts_headers = ['Date', 'Order Count', 'Revenue', 'Avg Order Value', 'Items Sold']
    ws_ts.append(ts_headers)

    # Build daily aggregation from orders
    order_dates = (
        Order.objects.filter(status__in=['pending', 'processing', 'shipped', 'delivered'])
        .values('created_at__date')
        .annotate(
            order_count=Count('id'),
            revenue=Sum('total_amount'),
        )
        .order_by('created_at__date')
    )
    for row in order_dates:
        dt = row['created_at__date']
        cnt = row['order_count']
        rev = float(row['revenue'] or 0)
        avg_val = round(rev / cnt, 2) if cnt else 0
        items = (
            OrderItem.objects.filter(order__created_at__date=dt)
            .aggregate(total_qty=Sum('quantity'))['total_qty'] or 0
        )
        ws_ts.append([_safe(dt) or 'N/A', cnt, rev, avg_val, items])
    _style_pbi_header(ws_ts, len(ts_headers))
    _auto_width(ws_ts)
    _make_table(ws_ts, 'PBI_SalesTimeSeries')

    # ── PBI-2. Revenue by Category ────────────────────────
    ws_cat = wb.create_sheet('PBI Revenue by Category')
    cat_rev_headers = ['Category', 'Revenue', 'Units Sold', 'Order Count', 'Avg Unit Price']
    ws_cat.append(cat_rev_headers)

    # Smart product order items (category is a char field on SmartProducts)
    from django.db.models.functions import Coalesce
    cat_stats = (
        OrderItem.objects
        .filter(smart_product__isnull=False)
        .values('smart_product__category')
        .annotate(
            revenue=Sum('subtotal'),
            units=Sum('quantity'),
            orders=Count('order', distinct=True),
        )
        .order_by('-revenue')
    )
    for cs in cat_stats:
        cat_name = cs['smart_product__category'] or 'Uncategorised'
        rev = float(cs['revenue'] or 0)
        units = cs['units'] or 0
        avg_price = round(rev / units, 2) if units else 0
        ws_cat.append([cat_name, rev, units, cs['orders'], avg_price])

    # Regular product order items (category via FK)
    cat_stats_reg = (
        OrderItem.objects
        .filter(product__isnull=False)
        .values('product__category__name')
        .annotate(
            revenue=Sum('subtotal'),
            units=Sum('quantity'),
            orders=Count('order', distinct=True),
        )
        .order_by('-revenue')
    )
    for cs in cat_stats_reg:
        cat_name = cs['product__category__name'] or 'Uncategorised'
        rev = float(cs['revenue'] or 0)
        units = cs['units'] or 0
        avg_price = round(rev / units, 2) if units else 0
        ws_cat.append([cat_name, rev, units, cs['orders'], avg_price])
    _style_pbi_header(ws_cat, len(cat_rev_headers))
    _auto_width(ws_cat)
    _make_table(ws_cat, 'PBI_RevenueByCategory')

    # ── PBI-3. Top Products Performance (Enhanced with Real-Time Data) ───────────────
    ws_top = wb.create_sheet('PBI Top Products')
    top_headers = ['Product Name', 'Category', 'All-Time Revenue', 'All-Time Units',
                   'Order Count', 'Avg Unit Price', 'Recent Weekly Sales', 'Recent Monthly Sales',
                   'Stock Qty', 'Stock Status', 'Demand Trend', 'Predicted Demand 7d', 'Predicted Demand 30d',
                   'Forecast Accuracy %', 'Performance Score']
    ws_top.append(top_headers)

    top_items = (
        OrderItem.objects
        .values('product_name')
        .annotate(
            revenue=Sum('subtotal'),
            units=Sum('quantity'),
            orders=Count('order', distinct=True),
            avg_price=Avg('unit_price'),
        )
        .order_by('-revenue')
    )
    
    enhanced_products = []
    for ti in top_items:
        # Get real-time sales data
        sales_data = _calculate_real_time_sales_data(ti['product_name'])
        
        # Try to find matching product for extra fields
        sp = SmartProducts.objects.filter(name=ti['product_name']).first()
        rp = Product.objects.select_related('category').filter(name=ti['product_name']).first()
        
        cat = 'N/A'
        stock = 0
        stock_status = 'Unknown'
        trend = 'N/A'
        pred7 = 0
        pred30 = 0
        forecast_acc = 0
        
        if sp:
            cat = _text(sp.category)
            stock = sp.stock_quantity or 0
            stock_info = _get_stock_status_info(sp)
            stock_status = stock_info['status']
            trend = _text(sp.demand_trend)
            pred7 = sp.predicted_demand_7d or 0
            pred30 = sp.predicted_demand_30d or 0
            forecast_acc = _safe(sp.forecast_accuracy) or 0
        elif rp:
            cat = _text(rp.category.name if rp.category else None)
            stock = rp.stock_quantity or 0
            stock_info = _get_stock_status_info(rp)
            stock_status = stock_info['status']
            trend = _text(rp.demand_trend)
            pred7 = rp.predicted_demand_7d or 0
            pred30 = rp.predicted_demand_30d or 0
            forecast_acc = _safe(rp.forecast_accuracy) or 0
        
        # Calculate performance score (revenue + recent activity)
        revenue_score = float(ti['revenue'] or 0) / 1000  # Scale down revenue
        recent_activity_score = (sales_data['avg_weekly_qty'] * 10) + (sales_data['avg_monthly_qty'] * 2)
        performance_score = round(revenue_score + recent_activity_score, 2)
        
        enhanced_products.append([
            _text(ti['product_name']), cat, float(ti['revenue'] or 0),
            ti['units'] or 0, ti['orders'] or 0, round(float(ti['avg_price'] or 0), 2),
            sales_data['avg_weekly_qty'], sales_data['avg_monthly_qty'],
            stock, stock_status, trend, pred7, pred30, forecast_acc, performance_score
        ])
    
    # Sort by performance score (highest first)
    enhanced_products.sort(key=lambda x: x[-1], reverse=True)
    
    for product_data in enhanced_products:
        ws_top.append(product_data)
    _style_pbi_header(ws_top, len(top_headers))
    _auto_width(ws_top)
    _make_table(ws_top, 'PBI_TopProducts')

    # ── PBI-4. Inventory Alerts (Enhanced with Real-Time Calculations) ───────────────────────
    ws_inv = wb.create_sheet('PBI Inventory Alerts')
    inv_headers = ['Product Name', 'Product Type', 'Category', 'Current Stock',
                   'Stock Status', 'Reorder Point', 'Min Stock', 'Max Stock',
                   'Recent Weekly Sales', 'Predicted Demand 7d', 'Days Until Stockout',
                   'Suggested Reorder Qty', 'Priority', 'Last Stock Movement', 'Demand Velocity']
    ws_inv.append(inv_headers)

    inventory_alerts = []
    
    # Regular products needing attention
    for p in Product.objects.select_related('category').filter(
        stock_quantity__isnull=False
    ).order_by('stock_quantity'):
        sales_data = _calculate_real_time_sales_data(p.name)
        stock_info = _get_stock_status_info(p)
        
        # Calculate demand velocity (units per day based on recent sales)
        demand_velocity = round(sales_data['avg_weekly_qty'] / 7, 2) if sales_data['avg_weekly_qty'] > 0 else 0
        
        # Calculate suggested reorder quantity
        current_stock = p.stock_quantity or 0
        max_stock = p.max_stock_level or 0
        suggested_reorder = max(0, max_stock - current_stock) if max_stock > 0 else p.reorder_point * 2
        
        # Get last stock movement
        last_movement = Inventory.objects.filter(
            product=p
        ).order_by('-change_date').first()
        last_movement_date = last_movement.change_date.strftime('%Y-%m-%d') if last_movement else 'Never'
        
        # Only include products that need attention
        if (current_stock <= p.reorder_point or 
            current_stock <= p.min_stock_level or 
            stock_info['status'] in ['Out of Stock', 'Critical Low', 'Needs Reorder']):
            
            inventory_alerts.append([
                _text(p.name), 'Regular', _text(p.category.name if p.category else None),
                current_stock, stock_info['status'], p.reorder_point or 0, 
                p.min_stock_level or 0, p.max_stock_level or 0,
                sales_data['avg_weekly_qty'], p.predicted_demand_7d or 0, 
                stock_info['days_of_stock'], suggested_reorder,
                'Critical' if current_stock <= p.min_stock_level else 'High' if current_stock <= p.reorder_point else 'Medium',
                last_movement_date, demand_velocity
            ])
    
    # Smart products needing attention  
    for p in SmartProducts.objects.filter(
        stock_quantity__isnull=False
    ).order_by('stock_quantity'):
        sales_data = _calculate_real_time_sales_data(p.name)
        stock_info = _get_stock_status_info(p)
        
        demand_velocity = round(sales_data['avg_weekly_qty'] / 7, 2) if sales_data['avg_weekly_qty'] > 0 else 0
        
        current_stock = p.stock_quantity or 0
        max_stock = p.max_stock_level or 0
        suggested_reorder = max(0, max_stock - current_stock) if max_stock > 0 else p.reorder_point * 2
        
        # Get last stock movement (Note: Inventory model may not link to SmartProducts directly)
        last_movement_date = 'N/A'  # Since SmartProducts may not have Inventory tracking
        
        if (current_stock <= p.reorder_point or 
            current_stock <= p.min_stock_level or 
            stock_info['status'] in ['Out of Stock', 'Critical Low', 'Needs Reorder']):
            
            inventory_alerts.append([
                _text(p.name), 'Smart', _text(p.category),
                current_stock, stock_info['status'], p.reorder_point or 0, 
                p.min_stock_level or 0, p.max_stock_level or 0,
                sales_data['avg_weekly_qty'], p.predicted_demand_7d or 0, 
                stock_info['days_of_stock'], suggested_reorder,
                'Critical' if current_stock <= p.min_stock_level else 'High' if current_stock <= p.reorder_point else 'Medium',
                last_movement_date, demand_velocity
            ])
    
    # Sort by priority and stock level
    priority_order = {'Critical': 1, 'High': 2, 'Medium': 3}
    inventory_alerts.sort(key=lambda x: (priority_order.get(x[12], 4), x[3]))
    
    for alert_data in inventory_alerts:
        ws_inv.append(alert_data)
    _style_pbi_header(ws_inv, len(inv_headers))
    _auto_width(ws_inv)
    _make_table(ws_inv, 'PBI_InventoryAlerts')

    # ── PBI-5. Demand Forecasting ─────────────────────────
    ws_fc = wb.create_sheet('PBI Demand Forecast')
    fc_headers = ['Product Name', 'Product Type', 'Category', 'Price',
                  'Current Stock', 'Peak Season', 'Festival',
                  'Demand Trend', 'Avg Weekly Sales', 'Avg Monthly Sales',
                  'Predicted Demand 7d', 'Predicted Demand 30d',
                  'Predicted Revenue 30d', 'Forecast Accuracy %',
                  'Needs Restock', 'Weather Dependent', 'Season Multiplier']
    ws_fc.append(fc_headers)

    for p in Product.objects.select_related('category').all():
        ws_fc.append([
            _text(p.name), 'Regular', _text(p.category.name if p.category else None),
            _safe(p.price) or 0, p.stock_quantity or 0, _text(p.peak_season), _text(p.festival_association),
            _text(p.demand_trend), _safe(p.avg_weekly_sales) or 0, _safe(p.avg_monthly_sales) or 0,
            p.predicted_demand_7d or 0, p.predicted_demand_30d or 0,
            _safe(p.predicted_revenue_30d) or 0, _safe(p.forecast_accuracy) or 0,
            'Yes' if p.needs_restock() else 'No',
            'Yes' if p.weather_dependent else 'No',
            p.get_current_season_multiplier(),
        ])
    for p in SmartProducts.objects.all():
        ws_fc.append([
            _text(p.name), 'Smart', _text(p.category),
            _safe(p.price) or 0, p.stock_quantity or 0, _text(p.peak_season), _text(p.festival_association),
            _text(p.demand_trend), _safe(p.avg_weekly_sales) or 0, _safe(p.avg_monthly_sales) or 0,
            p.predicted_demand_7d or 0, p.predicted_demand_30d or 0,
            _safe(p.predicted_revenue_30d) or 0, _safe(p.forecast_accuracy) or 0,
            'Yes' if p.needs_restock() else 'No',
            'Yes' if p.weather_dependent else 'No',
            p.get_current_season_multiplier(),
        ])
    _style_pbi_header(ws_fc, len(fc_headers))
    _auto_width(ws_fc)
    _make_table(ws_fc, 'PBI_DemandForecast')

    # ── PBI-6. Order Status Breakdown ─────────────────────
    ws_status = wb.create_sheet('PBI Order Status')
    status_headers = ['Status', 'Order Count', 'Total Revenue', 'Avg Order Value', '% of Total']
    ws_status.append(status_headers)
    total_ord_count = orders.count() or 1
    status_agg = (
        orders.values('status')
        .annotate(cnt=Count('id'), rev=Sum('total_amount'))
        .order_by('-rev')
    )
    for sa in status_agg:
        cnt = sa['cnt']
        rev = float(sa['rev'] or 0)
        avg_v = round(rev / cnt, 2) if cnt else 0
        pct = round(cnt / total_ord_count * 100, 1)
        ws_status.append([sa['status'], cnt, rev, avg_v, pct])
    _style_pbi_header(ws_status, len(status_headers))
    _auto_width(ws_status)
    _make_table(ws_status, 'PBI_OrderStatus')

    # ── PBI-7. Payment Analytics ──────────────────────────
    ws_pay_a = wb.create_sheet('PBI Payment Analytics')
    pa_headers = ['Payment Method', 'Count', 'Total Amount', 'Avg Amount',
                  'Completed', 'Failed', 'Success Rate %']
    ws_pay_a.append(pa_headers)
    pay_methods = (
        Payment.objects.values('payment_method')
        .annotate(
            cnt=Count('id'),
            total=Sum('amount'),
            avg_amt=Avg('amount'),
            completed=Count('id', filter=Q(status='completed')),
            failed=Count('id', filter=Q(status='failed')),
        )
        .order_by('-total')
    )
    for pm in pay_methods:
        cnt = pm['cnt'] or 1
        success = round((pm['completed'] or 0) / cnt * 100, 1)
        ws_pay_a.append([
            pm['payment_method'], pm['cnt'], float(pm['total'] or 0),
            round(float(pm['avg_amt'] or 0), 2), pm['completed'], pm['failed'], success,
        ])
    _style_pbi_header(ws_pay_a, len(pa_headers))
    _auto_width(ws_pay_a)
    _make_table(ws_pay_a, 'PBI_PaymentAnalytics')

    # ── PBI-8. ML Model Performance ───────────────────────
    ws_ml = wb.create_sheet('PBI ML Models')
    ml_headers = ['Model Name', 'Type', 'Forecast Type', 'Accuracy %',
                  'MAE', 'RMSE', 'MAPE', 'Active', 'Last Trained']
    ws_ml.append(ml_headers)
    for m in MLForecastModel.objects.all():
        ws_ml.append([
            _text(m.name), _text(m.model_type), _text(m.forecast_type), _safe(m.accuracy_score) or 0,
            _safe(m.mae) or 0, _safe(m.rmse) or 0, _safe(m.mape) or 0,
            'Yes' if m.is_active else 'No', _safe(m.last_trained) or 'N/A',
        ])
    _style_pbi_header(ws_ml, len(ml_headers))
    _auto_width(ws_ml)
    _make_table(ws_ml, 'PBI_MLModels')

    # ── PBI-9. KPI Summary ────────────────────────────────
    ws_kpi = wb.create_sheet('PBI KPIs')
    kpi_headers = ['KPI', 'Value']
    ws_kpi.append(kpi_headers)

    now = timezone.now()
    thirty_days_ago = now - timedelta(days=30)
    seven_days_ago = now - timedelta(days=7)

    orders_30d = orders.filter(created_at__gte=thirty_days_ago)
    orders_7d = orders.filter(created_at__gte=seven_days_ago)
    rev_30d = float(orders_30d.aggregate(s=Sum('total_amount'))['s'] or 0)
    rev_7d = float(orders_7d.aggregate(s=Sum('total_amount'))['s'] or 0)
    completed = orders.filter(status='delivered').count()
    completion_rate = round(completed / total_orders * 100, 1) if total_orders else 0
    avg_order = round(float(total_revenue) / total_orders, 2) if total_orders else 0

    active_users_30d = (User.objects.filter(last_login__gte=thirty_days_ago).count())

    smart_in_stock = SmartProducts.objects.filter(stock_quantity__gt=0).count()
    smart_total = SmartProducts.objects.count() or 1
    stock_rate = round(smart_in_stock / smart_total * 100, 1)

    kpi_rows = [
        ('Total Lifetime Revenue (Rs)', str(_safe(total_revenue))),
        ('Revenue Last 30 Days (Rs)', str(rev_30d)),
        ('Revenue Last 7 Days (Rs)', str(rev_7d)),
        ('Total Orders', str(total_orders)),
        ('Orders Last 30 Days', str(orders_30d.count())),
        ('Orders Last 7 Days', str(orders_7d.count())),
        ('Average Order Value (Rs)', str(avg_order)),
        ('Order Completion Rate %', str(completion_rate)),
        ('Total Products', str(total_products)),
        ('Products In Stock %', str(stock_rate)),
        ('Active Users (30d)', str(active_users_30d)),
        ('Total Registered Users', str(User.objects.count())),
        ('Total Customers', str(total_customers)),
        ('Pending Orders', str(orders.filter(status='pending').count())),
        ('Low Stock Alerts', str(SmartProducts.objects.filter(
            stock_quantity__isnull=False, stock_quantity__lte=models.F('reorder_point')
        ).count())),
        ('Unread Notifications', str(Notification.objects.filter(is_read=False).count())),
        ('Active ML Models', str(MLForecastModel.objects.filter(is_active=True).count())),
        ('Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]
    for row in kpi_rows:
        ws_kpi.append(list(row))
    _style_pbi_header(ws_kpi, 4)
    _auto_width(ws_kpi)
    _make_table(ws_kpi, 'PBI_KPIs')
    
    # ── NEW: Data Quality & Validation Sheet ───────────────────────────
    ws_quality = wb.create_sheet('Data Quality Report')
    quality_headers = ['Check Type', 'Item', 'Status', 'Count', 'Recommendation']
    ws_quality.append(quality_headers)
    
    quality_checks = [
        ('Stock Integrity', 'Products with Null Stock', 
         '⚠️ Issue' if Product.objects.filter(stock_quantity__isnull=True).exists() else '✅ Good',
         Product.objects.filter(stock_quantity__isnull=True).count(),
         'Update stock quantities for all products'),
        
        ('Stock Integrity', 'Smart Products with Null Stock',
         '⚠️ Issue' if SmartProducts.objects.filter(stock_quantity__isnull=True).exists() else '✅ Good', 
         SmartProducts.objects.filter(stock_quantity__isnull=True).count(),
         'Update stock quantities for all smart products'),
         
        ('ML Data Health', 'Products Never Forecasted',
         '⚠️ Issue' if outdated_forecasts > 0 else '✅ Good',
         outdated_forecasts,
         'Run ML forecasting for all products'),
         
        ('Sales Data', 'Products with Zero Sales History',
         '📊 Info',
         Product.objects.filter(avg_weekly_sales__lte=0).count(),
         'Monitor new products or discontinued items'),
         
        ('Pricing', 'Products with Zero Price',
         '⚠️ Issue' if Product.objects.filter(price__lte=0).exists() else '✅ Good',
         Product.objects.filter(price__lte=0).count(),
         'Update pricing for all products'),
         
        ('Inventory Alerts', 'Critical Stock Levels',
         '🚨 Urgent' if products_out_of_stock > 0 else '⚠️ Monitor' if products_low_stock > 0 else '✅ Good',
         products_out_of_stock + smart_products_out_of_stock,
         'Immediate restocking required'),
         
        ('Order Processing', 'Stuck Pending Orders',
         '⚠️ Review' if orders.filter(status='pending', created_at__lt=seven_days_ago).exists() else '✅ Good',
         orders.filter(status='pending', created_at__lt=seven_days_ago).count(),
         'Review and process old pending orders'),
         
        ('Customer Data', 'Orders without Customer Info',
         '📊 Info',
         Order.objects.filter(customer_email__isnull=True).count(),
         'Ensure customer data collection'),
         
        ('Revenue Tracking', 'Delivered Orders without Any Payments',
         '⚠️ Review' if Order.objects.filter(status='delivered').annotate(payment_count=Count('payments')).filter(payment_count=0).exists() else '✅ Good',
         Order.objects.filter(status='delivered').annotate(payment_count=Count('payments')).filter(payment_count=0).count(),
         'Link payments to all completed orders'),
         
        ('Data Freshness', 'Export Real-Time Status',
         '✅ Current',
         1, 
         'All calculations performed at export time')
    ]
    
    for check in quality_checks:
        ws_quality.append(list(check))
    
    _style_pbi_header(ws_quality, 5)
    _auto_width(ws_quality)
    _make_table(ws_quality, 'DataQualityReport')

    # ── Build response ────────────────────────────────────
    filename = f"business_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@login_required
def export_customer_excel(request):
    """Export the logged-in customer's own data to Excel for Power BI."""
    # Block owner / staff access
    if request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        messages.error(request, "This export is available for customer accounts only.")
        return redirect('smart_market:home')

    from .models import Order, OrderItem, Payment, Cart, CartItem

    wb = openpyxl.Workbook()

    # ── 1. Profile Summary ────────────────────────────────
    ws = wb.active
    ws.title = 'Profile Summary'
    ws.append(['Metric', 'Value'])

    customer_orders = Order.objects.filter(user=request.user).order_by('-created_at')
    total_orders = customer_orders.count()
    total_spent = float(customer_orders.aggregate(s=Sum('total_amount'))['s'] or 0)
    avg_order = (total_spent / total_orders) if total_orders > 0 else 0
    pending = customer_orders.filter(status='pending').count()
    completed = customer_orders.filter(status__in=['completed', 'delivered']).count()
    cancelled = customer_orders.filter(status='cancelled').count()
    loyalty = 'Premium' if total_orders > 5 else 'Standard'
    member_days = (datetime.now().date() - request.user.date_joined.date()).days

    summary_rows = [
        ('Customer Name', request.user.get_full_name() or request.user.username),
        ('Email', request.user.email),
        ('Member Since', request.user.date_joined.strftime('%Y-%m-%d')),
        ('Membership Days', str(member_days)),
        ('Loyalty Status', loyalty),
        ('Total Orders', str(total_orders)),
        ('Total Spent (Rs)', str(round(total_spent, 2))),
        ('Average Order Value', str(round(avg_order, 2))),
        ('Pending Orders', str(pending)),
        ('Completed Orders', str(completed)),
        ('Cancelled Orders', str(cancelled)),
        ('Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]
    for row in summary_rows:
        ws.append(list(row))
    _style_header_row(ws, 2)
    _auto_width(ws)
    _make_table(ws, 'CustomerProfile')

    # ── 2. Orders ─────────────────────────────────────────
    ws2 = wb.create_sheet('My Orders')
    ord_headers = ['Order #', 'Status', 'Subtotal', 'Tax', 'Shipping',
                   'Total', 'Shipping Address', 'City', 'Created']
    ws2.append(ord_headers)
    for o in customer_orders:
        ws2.append([
            o.order_number, _text(o.status),
            _safe(o.subtotal) or 0, _safe(o.tax_amount) or 0,
            _safe(o.shipping_cost) or 0, _safe(o.total_amount) or 0,
            _text(o.shipping_address), _text(o.shipping_city),
            _safe(o.created_at) or 'N/A',
        ])
    _style_header_row(ws2, len(ord_headers))
    _auto_width(ws2)
    _make_table(ws2, 'MyOrders')

    # ── 3. Order Items ────────────────────────────────────
    ws3 = wb.create_sheet('Order Items')
    oi_headers = ['Order #', 'Product Name', 'Unit Price', 'Quantity', 'Subtotal']
    ws3.append(oi_headers)
    for oi in OrderItem.objects.filter(order__user=request.user).select_related('order'):
        ws3.append([
            oi.order.order_number, _text(oi.product_name),
            _safe(oi.unit_price) or 0, oi.quantity or 0,
            _safe(oi.subtotal) or 0,
        ])
    _style_header_row(ws3, len(oi_headers))
    _auto_width(ws3)
    _make_table(ws3, 'MyOrderItems')

    # ── 4. Payments ───────────────────────────────────────
    ws4 = wb.create_sheet('Payments')
    pay_headers = ['Order #', 'Method', 'Amount', 'Status', 'Created']
    ws4.append(pay_headers)
    for pay in Payment.objects.filter(order__user=request.user).select_related('order'):
        ws4.append([
            pay.order.order_number, _text(pay.payment_method),
            _safe(pay.amount) or 0, _text(pay.status),
            _safe(pay.created_at) or 'N/A',
        ])
    _style_header_row(ws4, len(pay_headers))
    _auto_width(ws4)
    _make_table(ws4, 'MyPayments')

    # ── 5. Monthly Spending ───────────────────────────────
    ws5 = wb.create_sheet('Monthly Spending')
    spend_headers = ['Month', 'Orders', 'Total Spent']
    ws5.append(spend_headers)
    from datetime import timedelta
    for i in range(12):
        month_start = (datetime.now().replace(day=1) - timedelta(days=30 * i)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        monthly_qs = customer_orders.filter(
            created_at__date__gte=month_start.date(),
            created_at__date__lte=month_end.date()
        )
        monthly_total = float(monthly_qs.aggregate(s=Sum('total_amount'))['s'] or 0)
        ws5.append([month_start.strftime('%Y-%m'), monthly_qs.count(), round(monthly_total, 2)])
    _style_header_row(ws5, len(spend_headers))
    _auto_width(ws5)
    _make_table(ws5, 'MonthlySpending')

    # ── 6. Favorite Products ──────────────────────────────
    ws6 = wb.create_sheet('Favorite Products')
    fav_headers = ['Product Name', 'Times Purchased', 'Total Spent']
    ws6.append(fav_headers)
    favorites = (
        OrderItem.objects.filter(order__user=request.user)
        .values('product_name')
        .annotate(total_qty=Sum('quantity'), total_rev=Sum('subtotal'))
        .order_by('-total_qty')[:20]
    )
    for f in favorites:
        ws6.append([
            _text(f.get('product_name')),
            f.get('total_qty', 0),
            _safe(f.get('total_rev', 0)) or 0,
        ])
    _style_header_row(ws6, len(fav_headers))
    _auto_width(ws6)
    _make_table(ws6, 'FavoriteProducts')

    # ── Build response ────────────────────────────────────
    username = request.user.username
    filename = f"customer_dashboard_{username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


def about(request):
    # render a dedicated About page
    return render(request, 'about.html')


def add_to_cart(request, product_id):
    """Add product to cart - restricted for owners and staff"""
    # Prevent owners and staff from accessing cart functionality
    if request.user.is_authenticated and request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False, 
                'error': 'Cart functionality is not available for store owners and staff.'
            })
        messages.error(request, "Cart functionality is not available for store owners and staff.")
        return redirect('smart_market:shop')
    
    if request.method == 'POST':
        quantity = int(request.POST.get('quantity', 1))
        product_type = request.POST.get('product_type', 'smart')  # 'regular' or 'smart'
        
        try:
            # Get product to validate it exists and get name for message
            if product_type == 'smart':
                product = get_object_or_404(SmartProducts, id=product_id)
                available_stock = product.stock_quantity if product.stock_quantity is not None else 0
            else:
                product = get_object_or_404(Product, id=product_id)
                available_stock = product.stock_quantity if product.stock_quantity is not None else 0
                if hasattr(product, 'in_stock') and not product.in_stock:
                    available_stock = 0
            
            if available_stock <= 0:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': f"Sorry, {product.name} is out of stock."
                    })
                messages.error(request, f"Sorry, {product.name} is out of stock.")
                return redirect('smart_market:shop')
            
            if request.user.is_authenticated:
                # Authenticated user - use database cart
                cart, created = Cart.objects.get_or_create(user=request.user)
                
                if product_type == 'smart':
                    cart_item, item_created = CartItem.objects.get_or_create(
                        cart=cart, smart_product=product,
                        defaults={'quantity': quantity}
                    )
                else:
                    cart_item, item_created = CartItem.objects.get_or_create(
                        cart=cart, product=product,
                        defaults={'quantity': quantity}
                    )
                
                if not item_created:
                    cart_item.quantity += quantity
                else:
                    # item_created already has the quantity set via defaults
                    pass
                
                # Check total quantity doesn't exceed stock
                if cart_item.quantity > available_stock:
                    cart_item.quantity = available_stock
                    cart_item.save()
                    message = f"Only {available_stock} unit(s) of {product.name} available. Cart updated to maximum."
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': True,
                            'message': message,
                            'cart_count': cart.total_items,
                            'product_name': product.name,
                            'product_price': str(product.price),
                            'cart_total': str(cart.total_amount)
                        })
                    messages.warning(request, message)
                else:
                    cart_item.save()
                    message = f"Added {product.name} to cart!"
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': True,
                            'message': message,
                            'cart_count': cart.total_items,
                            'product_name': product.name,
                            'product_price': str(product.price),
                            'cart_total': str(cart.total_amount)
                        })
                    messages.success(request, message)
                return redirect('smart_market:cart')
                
            else:
                # Anonymous user - use session cart
                add_to_session_cart(request, product_id, product_type, quantity)
                message = f"Added {product.name} to cart! Login to checkout."
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    _, cart_count, cart_total = get_session_cart_items(request)
                    return JsonResponse({
                        'success': True,
                        'message': message,
                        'cart_count': cart_count,
                        'product_name': product.name,
                        'product_price': str(product.price),
                        'cart_total': str(cart_total)
                    })
                messages.success(request, message)
                return redirect('smart_market:cart')
            
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': f"Error adding product to cart: {str(e)}"
                })
            messages.error(request, f"Error adding product to cart: {str(e)}")
            return redirect('smart_market:shop')
    
    return redirect('smart_market:shop')


def cart(request):
    """Display shopping cart - restricted for owners and staff"""
    # Prevent owners and staff from accessing cart functionality
    if request.user.is_authenticated and request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        messages.error(request, "Cart functionality is not available for store owners and staff.")
        return redirect('smart_market:shop')
    
    if request.user.is_authenticated:
        # Authenticated user - get from database
        try:
            cart = Cart.objects.get(user=request.user)
            cart_items = cart.items.all()
            item_count = len(cart_items)
            subtotal = sum(item.subtotal for item in cart_items)
        except Cart.DoesNotExist:
            cart_items = []
            item_count = 0
            subtotal = Decimal('0.00')
    else:
        # Anonymous user - get from session
        session_items, item_count, subtotal = get_session_cart_items(request)
        cart_items = session_items
    
    # Calculate totals
    tax_rate = Decimal('0.15')  # 15% VAT
    tax_amount = subtotal * tax_rate
    shipping_cost = Decimal('50.00') if subtotal > 0 else Decimal('0.00')
    total = subtotal + tax_amount + shipping_cost
    
    context = {
        'cart_items': cart_items,
        'subtotal': subtotal,
        'tax_amount': tax_amount,
        'shipping_cost': shipping_cost,
        'total': total,
        'item_count': item_count,
        'is_guest': not request.user.is_authenticated,
        'guest_message': 'Please login or register to proceed to checkout' if not request.user.is_authenticated and item_count > 0 else None
    }
    
    return render(request, 'cart.html', context)


def cart_count(request):
    """Get cart count for AJAX requests - anonymous users and customers only"""
    # Owners and staff should not have access to cart functionality
    if request.user.is_authenticated and request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        return JsonResponse({'count': 0})
    
    if request.user.is_authenticated:
        try:
            cart = Cart.objects.get(user=request.user)
            count = cart.total_items
        except Cart.DoesNotExist:
            count = 0
    else:
        _, count, _ = get_session_cart_items(request)
    
    return JsonResponse({'count': count})


def mini_cart_data(request):
    """Get mini cart data for AJAX sidebar - anonymous users and customers only"""
    # Owners and staff should not have access to cart functionality
    if request.user.is_authenticated and request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        return JsonResponse({
            'success': False,
            'error': 'Cart functionality is not available for store owners and staff.'
        })
    
    try:
        if request.user.is_authenticated:
            try:
                cart = Cart.objects.get(user=request.user)
                cart_items = cart.items.all()
                items_data = []
                
                for item in cart_items:
                    items_data.append({
                        'id': item.id,
                        'product_id': item.product.id if item.product else item.smart_product.id,
                        'product_type': 'regular' if item.product else 'smart',
                        'name': item.product_name,
                        'price': str(item.unit_price),
                        'quantity': item.quantity,
                        'subtotal': str(item.subtotal),
                        'image_url': item.product.image_url if item.product else (item.smart_product.image_url if hasattr(item.smart_product, 'image_url') else None)
                    })
                
                return JsonResponse({
                    'success': True,
                    'items': items_data,
                    'count': cart.total_items,
                    'subtotal': str(cart.total_amount),
                    'total': str(cart.total_amount * Decimal('1.15') + Decimal('50.00'))  # 15% tax + shipping
                })
                
            except Cart.DoesNotExist:
                return JsonResponse({
                    'success': True,
                    'items': [],
                    'count': 0,
                    'subtotal': '0.00',
                    'total': '0.00'
                })
        else:
            # Anonymous user - get from session
            session_items, count, subtotal = get_session_cart_items(request)
            items_data = []
            
            for item in session_items:
                # Handle both regular and smart products
                if item['product_type'] == 'smart':
                    product_obj = item['smart_product']
                    product_id = item['smart_product'].id if item['smart_product'] else 0
                else:
                    product_obj = item['product']
                    product_id = item['product'].id if item['product'] else 0
                
                items_data.append({
                    'id': f"{item['product_type']}_{product_id}",
                    'product_id': product_id,
                    'product_type': item['product_type'],
                    'name': item.get('product_name', product_obj.name if product_obj else 'Unknown Product'),
                    'price': str(item.get('unit_price', product_obj.price if product_obj else 0)),
                    'quantity': item['quantity'],
                    'subtotal': str(item['subtotal']),
                    'image_url': product_obj.image_url if product_obj and hasattr(product_obj, 'image_url') else None
                })
            
            total = subtotal * Decimal('1.15') + (Decimal('50.00') if subtotal > 0 else Decimal('0.00'))
            
            return JsonResponse({
                'success': True,
                'items': items_data,
                'count': count,
                'subtotal': str(subtotal),
                'total': str(total)
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f"Error fetching cart data: {str(e)}"
        })


CHECKOUT_FORM_SESSION_KEY = 'checkout_form_data'


def checkout(request):
    """Checkout process for authenticated users - restricted for owners and staff"""
    # Require authentication for checkout
    if not request.user.is_authenticated:
        messages.info(request, "Please login or register to proceed with checkout.")
        return redirect('smart_market:unified_login')
    
    # Prevent owners and staff from accessing checkout functionality
    if request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        messages.error(request, "Checkout functionality is not available for store owners and staff.")
        return redirect('smart_market:shop')
    
    try:
        cart = Cart.objects.get(user=request.user)
        cart_items = cart.items.all()
    except Cart.DoesNotExist:
        messages.error(request, "Your cart is empty.")
        return redirect('smart_market:shop')
    
    if not cart_items:
        messages.error(request, "Your cart is empty.")
        return redirect('smart_market:shop')
    
    # Validate stock for all cart items before checkout
    stock_issues = []
    for item in cart_items:
        if item.smart_product:
            avail = item.smart_product.stock_quantity if item.smart_product.stock_quantity is not None else 0
            if avail <= 0:
                stock_issues.append(f'"{item.smart_product.name}" is out of stock and has been removed from your cart.')
                item.delete()
            elif item.quantity > avail:
                stock_issues.append(f'"{item.smart_product.name}" only has {avail} unit(s) available. Quantity adjusted.')
                item.quantity = avail
                item.save()
        elif item.product:
            avail = item.product.stock_quantity if item.product.stock_quantity is not None else 0
            out = hasattr(item.product, 'in_stock') and not item.product.in_stock
            if avail <= 0 or out:
                stock_issues.append(f'"{item.product.name}" is out of stock and has been removed from your cart.')
                item.delete()
            elif item.quantity > avail:
                stock_issues.append(f'"{item.product.name}" only has {avail} unit(s) available. Quantity adjusted.')
                item.quantity = avail
                item.save()
    
    if stock_issues:
        for issue in stock_issues:
            messages.warning(request, issue)
        # Re-check if cart is now empty
        cart_items = cart.items.all()
        if not cart_items.exists():
            messages.error(request, "All items in your cart are out of stock. Please add new items.")
            return redirect('smart_market:shop')
    
    # Calculate totals
    subtotal = cart.total_amount
    tax_rate = Decimal('0.15')
    tax_amount = subtotal * tax_rate
    default_shipping_cost = Decimal('50.00')
    cod_handling_fee = Decimal('25.00')
    shipping_cost = default_shipping_cost
    total = subtotal + tax_amount + shipping_cost
    
    if request.method == 'POST':
        # Process payment form
        return process_payment(request, cart, total)
    
    # Load user profile data first, then fall back to session data
    try:
        user_profile = request.user.profile
        profile_data = {
            'customer_phone': user_profile.phone or '',
            'delivery_method': user_profile.preferred_delivery_method or 'home_delivery',
            'shipping_address': user_profile.address or '',
            'shipping_city': user_profile.city or '',
            'shipping_postal_code': user_profile.postal_code or '',
            'pickup_store': user_profile.preferred_pickup_store or 'port_louis',
        }
    except UserProfile.DoesNotExist:
        profile_data = {}
    
    saved_checkout_data = request.session.get(CHECKOUT_FORM_SESSION_KEY, {})
    checkout_data = {
        'customer_name': saved_checkout_data.get('customer_name') or request.user.get_full_name(),
        'customer_email': saved_checkout_data.get('customer_email') or request.user.email,
        'customer_phone': saved_checkout_data.get('customer_phone') or profile_data.get('customer_phone', ''),
        'delivery_method': saved_checkout_data.get('delivery_method') or profile_data.get('delivery_method', 'home_delivery'),
        'shipping_address': saved_checkout_data.get('shipping_address') or profile_data.get('shipping_address', ''),
        'shipping_city': saved_checkout_data.get('shipping_city') or profile_data.get('shipping_city', ''),
        'shipping_postal_code': saved_checkout_data.get('shipping_postal_code') or profile_data.get('shipping_postal_code', ''),
        'pickup_store': saved_checkout_data.get('pickup_store') or profile_data.get('pickup_store', 'port_louis'),
        'save_details': saved_checkout_data.get('save_details', False),
        'payment_method': saved_checkout_data.get('payment_method', 'credit_card'),
    }
    
    context = {
        'cart_items': cart_items,
        'total_item_count': sum(item.quantity for item in cart_items),
        'subtotal': subtotal,
        'tax_amount': tax_amount,
        'shipping_cost': shipping_cost,
        'total': total,
        'default_shipping_cost': default_shipping_cost,
        'cod_handling_fee': cod_handling_fee,
        'checkout_data': checkout_data,
    }
    
    return render(request, 'checkout.html', context)


@login_required
def process_payment(request, cart=None, total_amount=None):
    """Process payment and create order"""
    try:
        if request.method != 'POST':
            return redirect('smart_market:checkout')

        # Get cart if not provided
        if not cart:
            cart = Cart.objects.get(user=request.user)
        
        # Extract form data
        payment_method = (request.POST.get('payment_method') or '').strip()
        customer_name = (request.POST.get('customer_name') or '').strip()
        customer_email = (request.POST.get('customer_email') or '').strip()
        customer_phone = (request.POST.get('customer_phone') or '').strip()
        delivery_method = (request.POST.get('delivery_method') or 'home_delivery').strip()
        
        # Address fields
        shipping_address = (request.POST.get('shipping_address') or '').strip()
        shipping_city = (request.POST.get('shipping_city') or '').strip()
        shipping_postal_code = (request.POST.get('shipping_postal_code') or '').strip()
        pickup_store = (request.POST.get('pickup_store') or '').strip()
        
        # Additional fields
        save_details = bool(request.POST.get('save_details'))
        accept_terms = request.POST.get('accept_terms')
        
        # Card info (for card payments)
        card_number = ''.join(char for char in request.POST.get('card_number', '') if char.isdigit())
        cardholder_name = (request.POST.get('cardholder_name') or '').strip()
        expiry = (request.POST.get('expiry') or '').strip()
        cvv = ''.join(char for char in request.POST.get('cvv', '') if char.isdigit())

        request.session[CHECKOUT_FORM_SESSION_KEY] = {
            'customer_name': customer_name,
            'customer_email': customer_email,
            'customer_phone': customer_phone,
            'delivery_method': delivery_method,
            'shipping_address': shipping_address,
            'shipping_city': shipping_city,
            'shipping_postal_code': shipping_postal_code,
            'pickup_store': pickup_store or 'port_louis',
            'save_details': save_details,
            'payment_method': payment_method or 'credit_card',
        }
        request.session.modified = True

        valid_payment_methods = {choice[0] for choice in Payment.PAYMENT_METHODS}
        if payment_method not in valid_payment_methods:
            messages.error(request, "Please select a valid payment method.")
            return redirect('smart_market:checkout')

        valid_delivery_methods = {'home_delivery', 'store_pickup'}
        if delivery_method not in valid_delivery_methods:
            messages.error(request, "Please select a valid delivery method.")
            return redirect('smart_market:checkout')

        store_addresses = {
            'port_louis': 'Smart Market - Port Louis, Royal Street, Port Louis',
            'rose_hill': 'Smart Market - Rose Hill, Vandermeersch Street, Rose Hill',
            'curepipe': 'Smart Market - Curepipe, Queen Mary Avenue, Curepipe'
        }
        
        # Validate required fields
        required_fields = [payment_method, customer_name, customer_email, customer_phone, accept_terms]
        
        # Add delivery-specific required fields
        if delivery_method == 'home_delivery':
            required_fields.extend([shipping_address, shipping_city])
        else:
            required_fields.append(pickup_store)

        if payment_method in ['credit_card', 'debit_card']:
            required_fields.extend([card_number, expiry, cvv, cardholder_name])
        
        if not all(required_fields):
            messages.error(request, "Please fill in all required fields and accept the terms.")
            return redirect('smart_market:checkout')

        # Format validation for fields
        import re
        name_pattern = re.compile(r"^[A-Za-z\s\-'.]{2,}$")
        phone_pattern = re.compile(r"^\+?[0-9\s\-]{7,15}$")

        if not name_pattern.match(customer_name):
            messages.error(request, "Name must contain only letters (no numbers or special characters).")
            return redirect('smart_market:checkout')

        if not phone_pattern.match(customer_phone):
            messages.error(request, "Phone number must contain only digits (e.g. +230 5XXX XXXX).")
            return redirect('smart_market:checkout')

        if delivery_method == 'home_delivery':
            if not re.search(r'[A-Za-z]', shipping_address):
                messages.error(request, "Address must include a street name (not just numbers).")
                return redirect('smart_market:checkout')
            if len(shipping_address) < 5:
                messages.error(request, "Please enter a complete street address.")
                return redirect('smart_market:checkout')

        if shipping_postal_code and not re.match(r'^[0-9]{3,6}$', shipping_postal_code):
            messages.error(request, "Postal code must be 3-6 digits only.")
            return redirect('smart_market:checkout')

        if payment_method in ['credit_card', 'debit_card']:
            if not name_pattern.match(cardholder_name):
                messages.error(request, "Cardholder name must contain only letters.")
                return redirect('smart_market:checkout')
            if not re.match(r'^[0-9]{13,19}$', card_number):
                messages.error(request, "Please enter a valid card number.")
                return redirect('smart_market:checkout')

        if delivery_method == 'store_pickup' and pickup_store not in store_addresses:
            messages.error(request, "Please select a valid pickup store.")
            return redirect('smart_market:checkout')
        
        # Calculate order totals based on delivery method
        subtotal = cart.total_amount
        tax_amount = subtotal * Decimal('0.15')
        
        # Calculate shipping cost
        if delivery_method == 'home_delivery':
            shipping_cost = Decimal('50.00')
        else:  # store_pickup
            shipping_cost = Decimal('0.00')
            
        # Add COD handling fee
        if payment_method == 'cash_on_delivery':
            shipping_cost += Decimal('25.00')
            
        total = subtotal + tax_amount + shipping_cost
        
        # Set delivery address based on method
        if delivery_method == 'store_pickup':
            shipping_address = store_addresses.get(pickup_store, 'Store Pickup')
            shipping_city = pickup_store.replace('_', ' ').title()
        
        # Create order
        order = Order.objects.create(
            user=request.user,
            customer_name=customer_name,
            customer_email=customer_email,
            customer_phone=customer_phone,
            delivery_method=delivery_method,
            pickup_store=pickup_store if delivery_method == 'store_pickup' else None,
            shipping_address=shipping_address,
            shipping_city=shipping_city,
            shipping_postal_code=shipping_postal_code,
            subtotal=subtotal,
            tax_amount=tax_amount,
            shipping_cost=shipping_cost,
            total_amount=total
        )
        
        # Create order items - optimized with batch notifications
        stock_notifications = []
        owner_users = list(User.objects.filter(groups__name='Owner').distinct())
        cart_items_qs = cart.items.select_related('product', 'smart_product').all()
        
        for cart_item in cart_items_qs:
            OrderItem.objects.create(
                order=order,
                product=cart_item.product,
                smart_product=cart_item.smart_product,
                product_name=cart_item.product_name,
                unit_price=cart_item.unit_price,
                quantity=cart_item.quantity,
                subtotal=cart_item.subtotal
            )

            # Decrement stock quantity after order
            if cart_item.product and cart_item.product.stock_quantity is not None:
                cart_item.product.stock_quantity = max(0, cart_item.product.stock_quantity - cart_item.quantity)
                if cart_item.product.stock_quantity == 0:
                    cart_item.product.in_stock = False
                cart_item.product.save()

                if cart_item.product.stock_quantity == 0:
                    for owner in owner_users:
                        stock_notifications.append(Notification(
                            recipient_user=owner,
                            notification_type='low_stock',
                            title=f'Out of Stock: {cart_item.product.name}',
                            message=f'"{cart_item.product.name}" is now OUT OF STOCK after Order #{order.order_number}.',
                            related_order=order
                        ))
                elif cart_item.product.stock_quantity <= cart_item.product.reorder_point:
                    for owner in owner_users:
                        stock_notifications.append(Notification(
                            recipient_user=owner,
                            notification_type='low_stock',
                            title=f'Low Stock: {cart_item.product.name}',
                            message=f'"{cart_item.product.name}" — only {cart_item.product.stock_quantity} left. Order #{order.order_number}.',
                            related_order=order
                        ))

            elif cart_item.smart_product and cart_item.smart_product.stock_quantity is not None:
                cart_item.smart_product.stock_quantity = max(0, cart_item.smart_product.stock_quantity - cart_item.quantity)
                cart_item.smart_product.save()

                if cart_item.smart_product.stock_quantity == 0:
                    for owner in owner_users:
                        stock_notifications.append(Notification(
                            recipient_user=owner,
                            notification_type='low_stock',
                            title=f'Out of Stock: {cart_item.smart_product.name}',
                            message=f'"{cart_item.smart_product.name}" is now OUT OF STOCK after Order #{order.order_number}.',
                            related_order=order
                        ))
                elif cart_item.smart_product.stock_quantity <= cart_item.smart_product.reorder_point:
                    for owner in owner_users:
                        stock_notifications.append(Notification(
                            recipient_user=owner,
                            notification_type='low_stock',
                            title=f'Low Stock: {cart_item.smart_product.name}',
                            message=f'"{cart_item.smart_product.name}" — only {cart_item.smart_product.stock_quantity} left. Order #{order.order_number}.',
                            related_order=order
                        ))
        
        # Batch create all stock notifications in one query
        if stock_notifications:
            Notification.objects.bulk_create(stock_notifications)
        
        # Create payment record
        payment = Payment.objects.create(
            order=order,
            payment_method=payment_method,
            amount=total,
            billing_name=cardholder_name or customer_name,
            billing_address=shipping_address,
            billing_city=shipping_city,
            billing_postal_code=shipping_postal_code,
            transaction_id=f"TXN-{order.order_number}"
        )
        
        # Set card details if card payment
        if payment_method in ['credit_card', 'debit_card'] and card_number:
            payment.card_last_four = card_number[-4:] if len(card_number) >= 4 else ''
            payment.card_brand = 'Unknown'  # In real app, detect card brand
            
        # Simulate payment processing
        if payment_method == 'cash_on_delivery':
            payment.status = 'pending'
            order.status = 'processing'
        elif payment_method in ['mobile_money', 'bank_transfer']:
            payment.status = 'pending' 
            order.status = 'pending'
        else:
            payment.status = 'completed'
            order.status = 'processing'
            
        payment.save()
        order.save()
        
        # Create new order notification for owners (batch)
        order_notifs = []
        item_names = ", ".join(item.product_name for item in order.items.all())
        for owner in owner_users:
            order_notifs.append(Notification(
                recipient_user=owner,
                notification_type='new_order',
                title=f'New Order #{order.order_number} by {customer_name}',
                message=f'{customer_name} placed a new order for Rs {total:.2f}. Payment: {payment.get_payment_method_display()}. Items: {item_names}.',
                related_order=order
            ))
        if order_notifs:
            Notification.objects.bulk_create(order_notifs)

        if save_details:
            # Save to session for form persistence
            request.session[CHECKOUT_FORM_SESSION_KEY] = {
                'customer_name': customer_name,
                'customer_email': customer_email,
                'customer_phone': customer_phone,
                'delivery_method': delivery_method,
                'shipping_address': shipping_address if delivery_method == 'home_delivery' else '',
                'shipping_city': shipping_city if delivery_method == 'home_delivery' else '',
                'shipping_postal_code': shipping_postal_code if delivery_method == 'home_delivery' else '',
                'pickup_store': pickup_store or 'port_louis',
                'save_details': True,
                'payment_method': payment_method,
            }
            
            # Save to user profile for persistence across sessions
            try:
                user_profile = request.user.profile
            except UserProfile.DoesNotExist:
                user_profile = UserProfile.objects.create(user=request.user)
            
            user_profile.phone = customer_phone
            user_profile.address = shipping_address if delivery_method == 'home_delivery' else ''
            user_profile.city = shipping_city if delivery_method == 'home_delivery' else ''
            user_profile.postal_code = shipping_postal_code if delivery_method == 'home_delivery' else ''
            user_profile.preferred_delivery_method = delivery_method
            user_profile.preferred_pickup_store = pickup_store or 'port_louis'
            user_profile.save()
        else:
            request.session.pop(CHECKOUT_FORM_SESSION_KEY, None)
        request.session.modified = True
        
        # Clear cart after successful order
        cart.delete()
        
        messages.success(request, f"Order #{order.order_number} placed successfully!")
        return redirect('smart_market:order_confirmation', order_number=order.order_number)
        
    except Exception as e:
        messages.error(request, f"Payment processing failed: {str(e)}")
        return redirect('smart_market:checkout')


def remove_from_cart(request, product_id):
    """Remove product from cart (supports both authenticated and anonymous users) - restricted for owners and staff"""
    # Prevent owners and staff from accessing cart functionality
    if request.user.is_authenticated and request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        messages.error(request, "Cart functionality is not available for store owners and staff.")
        return redirect('smart_market:shop')
    
    if request.user.is_authenticated:
        # Authenticated user - remove from database
        try:
            cart = Cart.objects.get(user=request.user)
            product_type = request.POST.get('type') or request.GET.get('type', 'smart')
            
            if product_type == 'smart':
                CartItem.objects.filter(cart=cart, smart_product_id=product_id).delete()
            else:
                CartItem.objects.filter(cart=cart, product_id=product_id).delete()
                
            messages.success(request, "Item removed from cart!")
        except Cart.DoesNotExist:
            pass
    else:
        # Anonymous user - remove from session
        cart = get_session_cart(request)
        product_type = request.POST.get('type') or request.GET.get('type', 'smart')
        item_key = f"{product_type}_{product_id}"
        
        if item_key in cart:
            del cart[item_key]
            save_session_cart(request, cart)
            messages.success(request, "Item removed from cart!")
    
    return redirect('smart_market:cart')


def update_cart(request, product_id):
    """Update product quantity in cart (supports both authenticated and anonymous users) - restricted for owners and staff"""
    # Prevent owners and staff from accessing cart functionality
    if request.user.is_authenticated and request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Cart functionality not available for owners/staff'})
        messages.error(request, "Cart functionality is not available for store owners and staff.")
        return redirect('smart_market:shop')
    
    if request.method == 'POST':
        quantity = int(request.POST.get('quantity', 1))
        
        if quantity <= 0:
            return remove_from_cart(request, product_id)
        
        if request.user.is_authenticated:
            # Authenticated user - update database
            try:
                cart = Cart.objects.get(user=request.user)
                product_type = request.POST.get('product_type', 'smart')
                
                if product_type == 'smart':
                    cart_item = CartItem.objects.get(cart=cart, smart_product_id=product_id)
                    available_stock = cart_item.smart_product.stock_quantity if cart_item.smart_product.stock_quantity is not None else 0
                else:
                    cart_item = CartItem.objects.get(cart=cart, product_id=product_id)
                    available_stock = cart_item.product.stock_quantity if cart_item.product.stock_quantity is not None else 0
                
                if quantity > available_stock:
                    quantity = available_stock
                    messages.warning(request, f"Only {available_stock} unit(s) available. Quantity adjusted.")
                
                cart_item.quantity = quantity
                cart_item.save()
                
                # AJAX response for better UX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    cart_summary = {
                        'item_total': float(cart_item.subtotal),
                        'cart_total': float(cart.total_amount),
                        'item_count': cart.total_items,
                        'success': True
                    }
                    return JsonResponse(cart_summary)
                
                messages.success(request, "Cart updated!")
            except (Cart.DoesNotExist, CartItem.DoesNotExist):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': 'Item not found in cart'})
                messages.error(request, "Item not found in cart!")
        else:
            # Anonymous user - update session
            cart = get_session_cart(request)
            product_type = request.POST.get('product_type', 'smart')
            item_key = f"{product_type}_{product_id}"
            
            if item_key in cart:
                cart[item_key]['quantity'] = quantity
                save_session_cart(request, cart)
                
                # AJAX response for guest users
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    _, item_count, total_amount = get_session_cart_items(request)
                    cart_summary = {
                        'cart_total': float(total_amount),
                        'item_count': item_count,
                        'success': True
                    }
                    return JsonResponse(cart_summary)
                    
                messages.success(request, "Cart updated!")
    
    return redirect('smart_market:cart')


def clear_cart(request):
    """Clear all items from cart - restricted for owners and staff"""
    # Prevent owners and staff from accessing cart functionality
    if request.user.is_authenticated and request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        messages.error(request, "Cart functionality is not available for store owners and staff.")
        return redirect('smart_market:shop')
    
    if request.user.is_authenticated:
        # Authenticated user - clear database cart
        try:
            cart = Cart.objects.get(user=request.user)
            cart.items.all().delete()
            messages.success(request, "Cart cleared successfully!")
        except Cart.DoesNotExist:
            messages.info(request, "Cart is already empty.")
    else:
        # Anonymous user - clear session cart
        clear_session_cart(request)
        messages.success(request, "Cart cleared successfully!")
    
    return redirect('smart_market:cart')


@login_required
def mark_notification_read(request, notification_id):
    """Mark notification as read"""
    try:
        notification = Notification.objects.get(id=notification_id, recipient_user=request.user)
        notification.is_read = True
        notification.save()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
            
    except Notification.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Notification not found'})
    
    return redirect(request.META.get('HTTP_REFERER', 'smart_market:owner_dashboard'))


@login_required 
def mark_all_notifications_read(request):
    """Mark all notifications as read for current user"""
    Notification.objects.filter(recipient_user=request.user, is_read=False).update(is_read=True)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
        
    messages.success(request, "All notifications marked as read.")
    return redirect(request.META.get('HTTP_REFERER', 'smart_market:owner_dashboard'))


def generate_password_reset_token(email):
    serializer = URLSafeTimedSerializer(settings.SECRET_KEY)
    return serializer.dumps(email, salt='password-reset-salt')

def verify_password_reset_token(token, max_age=3600):
    serializer = URLSafeTimedSerializer(settings.SECRET_KEY)
    try:
        email = serializer.loads(token, salt='password-reset-salt', max_age=max_age)
        return email
    except (BadSignature, SignatureExpired):
        return None

def forgot_password(request):
    if request.method == 'POST':
        action = request.POST.get('action', 'send_code')

        if action == 'send_code':
            email = request.POST.get('email', '').strip()
            user = User.objects.filter(email=email).first()
            if user:
                otp_code = str(random.randint(100000, 999999))
                request.session['reset_otp'] = otp_code
                request.session['reset_email'] = email
                send_mail(
                    'Password Reset Code',
                    f'Your password reset code is: {otp_code}',
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=True,
                )
                return render(request, 'forgot_password.html', {
                    'step': 'verify',
                    'email': email,
                    'otp_code': otp_code,
                })
            return render(request, 'forgot_password.html', {
                'error': 'No account found with that email address.'
            })

        elif action == 'verify_code':
            entered_code = request.POST.get('otp_code', '').strip()
            stored_code = request.session.get('reset_otp')
            stored_email = request.session.get('reset_email')
            if stored_code and entered_code == stored_code and stored_email:
                token = generate_password_reset_token(stored_email)
                del request.session['reset_otp']
                del request.session['reset_email']
                return redirect('smart_market:reset_password', token=token)
            return render(request, 'forgot_password.html', {
                'step': 'verify',
                'email': stored_email or '',
                'error': 'Invalid code. Please try again.',
            })

    return render(request, 'forgot_password.html')


def reset_password(request, token):
    email = verify_password_reset_token(token)
    if not email:
        return render(request, 'reset_password.html', {'error': 'Invalid or expired reset link.'})
    if request.method == 'POST':
        password = request.POST.get('password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()
        if not password or not confirm_password:
            return render(request, 'reset_password.html', {'error': 'All fields are required.'})
        if password != confirm_password:
            return render(request, 'reset_password.html', {'error': 'Passwords do not match.'})
        user = User.objects.filter(email=email).first()
        if user:
            user.set_password(password)
            user.save()
            return render(request, 'reset_password.html', {'message': 'Your password has been reset successfully.'})
        return render(request, 'reset_password.html', {'error': 'User not found.'})
    return render(request, 'reset_password.html')


# Enhanced Cart and Order Management Views

@login_required
@login_required
def order_confirmation(request, order_number):
    """Display order confirmation"""
    order = get_object_or_404(Order, order_number=order_number, user=request.user)
    
    context = {
        'order': order,
        'order_items': order.items.all(),
        'payment': order.payments.first(),
    }
    
    return render(request, 'order_confirmation.html', context)


@login_required
def my_orders(request):
    """Display user's order history"""
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    
    context = {
        'orders': orders,
    }
    
    return render(request, 'my_orders.html', context)


@login_required
@login_required
def order_detail(request, order_number):
    """Display detailed order information"""
    # Check if user is an owner (can view any order) or customer (can only view their own)
    if request.user.groups.filter(name='Owner').exists():
        # Owners can view any order
        order = get_object_or_404(Order, order_number=order_number)
    else:
        # Regular customers can only view their own orders
        order = get_object_or_404(Order, order_number=order_number, user=request.user)
    
    context = {
        'order': order,
        'order_items': order.items.all(),
        'payments': order.payments.all(),
        'is_owner_view': request.user.groups.filter(name='Owner').exists(),
    }
    
    return render(request, 'order_detail.html', context)


@login_required
@require_POST
def update_cart_item(request, item_id):
    """Update cart item quantity - restricted for owners and staff"""
    # Prevent owners and staff from accessing cart functionality
    if request.user.is_authenticated and request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        messages.error(request, "Cart functionality is not available for store owners and staff.")
        return redirect('smart_market:shop')
    
    try:
        cart = Cart.objects.get(user=request.user)
        cart_item = get_object_or_404(CartItem, id=item_id, cart=cart)
        
        new_quantity = int(request.POST.get('quantity', 1))
        
        if new_quantity <= 0:
            cart_item.delete()
            messages.success(request, "Item removed from cart.")
        else:
            # Enforce stock limit
            if cart_item.smart_product:
                available_stock = cart_item.smart_product.stock_quantity if cart_item.smart_product.stock_quantity is not None else 0
            elif cart_item.product:
                available_stock = cart_item.product.stock_quantity if cart_item.product.stock_quantity is not None else 0
            else:
                available_stock = new_quantity
            
            if new_quantity > available_stock:
                new_quantity = available_stock
                messages.warning(request, f"Only {available_stock} unit(s) available. Quantity adjusted.")
            else:
                messages.success(request, "Cart updated successfully.")
            
            cart_item.quantity = new_quantity
            cart_item.save()
            
    except Cart.DoesNotExist:
        messages.error(request, "Cart not found.")
    except ValueError:
        messages.error(request, "Invalid quantity.")
    
    return redirect('smart_market:cart')


@login_required
@require_POST
def remove_cart_item(request, item_id):
    """Remove item from cart - restricted for owners and staff"""
    # Prevent owners and staff from accessing cart functionality
    if request.user.groups.filter(name__in=['Owner', 'Staff']).exists():
        messages.error(request, "Cart functionality is not available for store owners and staff.")
        return redirect('smart_market:shop')
    
    try:
        cart = Cart.objects.get(user=request.user)
        cart_item = get_object_or_404(CartItem, id=item_id, cart=cart)
        
        product_name = cart_item.product_name
        cart_item.delete()
        
        messages.success(request, f"Removed {product_name} from cart.")
        
    except Cart.DoesNotExist:
        messages.error(request, "Cart not found.")
    
    return redirect('smart_market:cart')
